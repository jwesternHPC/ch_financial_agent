import os
import threading
import requests
import pandas as pd
from datetime import datetime, timedelta
from lxml import etree
import logging
import re
import numpy as np
import time
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
from copy import copy
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None

logging.getLogger('pypdf').setLevel(logging.ERROR)
logger = logging.getLogger(__name__)

# Companies House API base URL
BASE_URL = "https://api.company-information.service.gov.uk"

# XBRL tag mappings for UK FRS accounts
XBRL_TAG_MAPPINGS = {
    'Revenue': ['turnovergrossoperatingrevenue', 'turnover', 'sales', 'revenue', 'revenues'],
    'CostOfSales': ['costofsales', 'costofsalesandgrossexpense', 'costofrevenue', 'costs'],
    'GrossProfit': ['grossprofit', 'grossprofitloss'],
    'OperatingIncome': ['operatingprofitloss', 'operatingprofit', 'profitbeforetax', 'earningsbeforeinterestandtax', 'ebit'],
    'NetIncome': ['profitloss', 'profitlossforperiod', 'profitlossaftertax', 'netincome', 'profit'],
    'TotalAssets': ['balancesheettotal', 'totalassets'],
    'CurrentAssets': ['currentassets', 'currentassetslessstock', 'currentassetsincluding'],
    'FixedAssets': ['fixedassets', 'propertyplantequipment', 'intangibleassets'],
    'Equity': ['shareholdersequity', 'netassets', 'netassetsliabilities', 'equity'],
    'CurrentLiabilities': ['creditorsduewithinoneyear', 'currentliabilities', 'creditors', 'tradecreditorstradereceivables', 'tradecreditors', 'tradepayables'],
    'LongTermLiabilities': ['creditorsdueafteroneyear', 'longtermliabilities', 'longtermdebt', 'longtermborrowings', 'longtermloans'],
    'Inventory': ['inventory', 'stock'],
    'Cash': ['cashandcashequivalents', 'cashbankonhand', 'cashcashequivalents', 'cash'],
    'Receivables': ['tradedebtorstradereceivables', 'tradecreditors', 'receivables', 'debtors'],
    'DepreciationAmortization': ['depreciationandamortisation', 'depreciation', 'amortisation'],
    'Exceptionals': ['exceptionalitems', 'exceptionalcosts', 'exceptionalitem', 'exceptionaloperatingcosts', 'nonunderlyingitems'],
    'EBITDA': ['ebitda', 'earningsbeforeinterestandtaxdepreciationamortisation'],
    'Interest': ['interestexpense', 'interestpaid', 'financecosts', 'financecost', 'interest'],
    'Tax': ['taxexpense', 'incomeandcapitalgains', 'tax'],
    'Employees': [
        'consolidatedaveragenumberemployeesduringperiod',
        'groupaveragenumberemployeesduringperiod',
        'consolidatedaveragenumberemployees',
        'groupaveragenumberemployees',
        'averagenumberemployees',
        'averagenumberemployeesduringperiod',
    ],
    'EPS': ['earningspershare', 'eps', 'earningspersh'],
    'InterestIncome': ['interestincome'],
}

STATEMENT_TYPE_KEYWORDS = {
    'income_statement': set(
        XBRL_TAG_MAPPINGS['Revenue']
        + XBRL_TAG_MAPPINGS['CostOfSales']
        + XBRL_TAG_MAPPINGS['GrossProfit']
        + XBRL_TAG_MAPPINGS['OperatingIncome']
        + XBRL_TAG_MAPPINGS['NetIncome']
        + XBRL_TAG_MAPPINGS['DepreciationAmortization']
        + XBRL_TAG_MAPPINGS['EBITDA']
        + XBRL_TAG_MAPPINGS['Interest']
        + XBRL_TAG_MAPPINGS['Tax']
        + XBRL_TAG_MAPPINGS['EPS']
        + XBRL_TAG_MAPPINGS['InterestIncome']
    ),
    'balance_sheet': set(
        XBRL_TAG_MAPPINGS['TotalAssets']
        + XBRL_TAG_MAPPINGS['CurrentAssets']
        + XBRL_TAG_MAPPINGS['FixedAssets']
        + XBRL_TAG_MAPPINGS['Equity']
        + XBRL_TAG_MAPPINGS['CurrentLiabilities']
        + XBRL_TAG_MAPPINGS['LongTermLiabilities']
        + XBRL_TAG_MAPPINGS['Inventory']
        + XBRL_TAG_MAPPINGS['Cash']
        + XBRL_TAG_MAPPINGS['Receivables']
    ),
    'cash_flow': {
        'operatingcashflow', 'cashfromoperations', 'netcashfromoperatingactivities',
        'cashgeneratedfromoperations', 'netcashgeneratedfromoperatingactivities',
        'netcashprovidedbyoperatingactivities', 'capitalexpenditure', 'capex',
        'additionsotherthanthroughbusinesscombinationspropertyplantequipment',
        'additionsotherthanthroughbusinesscombinationsintangibleassets'
    },
}

TEMPLATE_FILE = "Excel Output Template v1.xlsx"
TEMPLATE_SHEET_NAME = "Sheet1"
TEMPLATE_START_COL = 6
TEMPLATE_MAX_YEARS = 10
TEMPLATE_ROW_MAP = {
    'Revenue': 5,
    'Revenue Growth (%)': 6,
    'Gross Profit': 8,
    'Gross Margin (%)': 9,
    'EBITDA': 11,
    'EBITDA Margin (%)': 12,
    'Exceptionals': 14,
    'Adjusted EBITDA': 15,
    'Depreciation and Amortisation': 17,
    'EBIT': 18,
    'EBIT Margin (%)': 19,
    'Working Capital Movement': 21,
    'Capital Expenditures': 22,
    'Tax Paid': 23,
    'Cash Flow from Operations': 24,
    'Net Cash Flow': 26,
    'Receivables Turnover': 28,
    'Days Sales Outstanding (DSO)': 29,
    'Inventory Turnover': 30,
    'Total Debt': 32,
    'Total Cash': 33,
    'Net Debt': 34,
    'Number of Employees': 36,
    'Notes:': 38,
}

MONEY_TEMPLATE_ROWS = {
    'Revenue', 'Gross Profit', 'EBITDA', 'Exceptionals', 'Adjusted EBITDA', 'Depreciation and Amortisation', 'EBIT',
    'Working Capital Movement', 'Capital Expenditures', 'Tax Paid',
    'Cash Flow from Operations', 'Net Cash Flow', 'Total Debt', 'Total Cash', 'Net Debt'
}

PERCENT_TEMPLATE_ROWS = {'Revenue Growth (%)', 'Gross Margin (%)', 'EBITDA Margin (%)', 'EBIT Margin (%)'}

COUNT_TEMPLATE_ROWS = {'Number of Employees'}

WHITE_FILL = PatternFill(fill_type='solid', start_color='FFFFFFFF', end_color='FFFFFFFF')
TEMPLATE_BLUE_FILLS = {'FF002060', '00002060'}

# ---------------------------------------------------------------------------
# Confidence score thresholds and limits
# ---------------------------------------------------------------------------
CONFIDENCE_HIGH_THRESHOLD = 85    # score >= this → 'high'
CONFIDENCE_MEDIUM_THRESHOLD = 65  # score >= this → 'medium'
CONFIDENCE_LOW_THRESHOLD = 45     # score >= this → 'low'; below → 'very low'
CONFIDENCE_MINIMUM = 35           # floor applied after all adjustments
CONFIDENCE_MAXIMUM = 95           # ceiling applied after all adjustments

# Fact-selection score → confidence conversion
CONFIDENCE_FACT_BASE = 45         # base confidence before selection-score contribution
CONFIDENCE_SCORE_CAP = 180        # selection scores are clamped to this before scaling
CONFIDENCE_SCORE_WEIGHT = 0.25    # fraction of selection score added to base confidence
CONFIDENCE_NO_DIMENSION_BONUS = 8 # bonus for non-dimensional (entity-wide) contexts
CONFIDENCE_DIMENSION_PENALTY = 5  # per-dimension confidence penalty
CONFIDENCE_DIMENSION_PENALTY_CAP = 20  # maximum total dimension penalty

# PDF fallback confidence scores (lower than XBRL; reflects unstructured-text uncertainty)
CONFIDENCE_PDF_NARRATIVE = 42     # narrative regex match from PDF body text
CONFIDENCE_PDF_NOTE = 48          # note disclosure parsed from PDF

# ---------------------------------------------------------------------------
# Context selection scoring weights
# ---------------------------------------------------------------------------
SCORE_DATE_PROXIMITY_MAX = 120    # day-difference beyond which proximity score is zero
SCORE_DATE_EXACT_BONUS = 100      # bonus when context date exactly matches filing date
SCORE_DURATION_CONTEXT_BONUS = 55 # income/cash-flow: prefer duration (start→end) contexts
SCORE_INSTANT_CONTEXT_PENALTY = 45  # income/cash-flow: penalise instant contexts
SCORE_INSTANT_CONTEXT_BONUS = 55  # balance-sheet: prefer instant contexts
SCORE_DURATION_CONTEXT_PENALTY = 35  # balance-sheet: penalise duration contexts
SCORE_UNKNOWN_DURATION_BONUS = 25 # unknown statement type: mild bonus for duration contexts
SCORE_UNKNOWN_INSTANT_BONUS = 20  # unknown statement type: mild bonus for instant contexts
SCORE_NO_DIMENSION_BONUS = 40     # bonus for non-dimensional (consolidated) contexts
SCORE_DIMENSION_PENALTY = 15      # penalty per dimension in segment/dimensional contexts

METRIC_SUMMARY_ORDER = [
    'Revenue',
    'Gross Profit',
    'Operating Income (EBIT)',
    'EBITDA',
    'Exceptionals',
    'Adjusted EBITDA',
    'Depreciation and Amortisation',
    'Net Income',
    'Cash Flow from Operations (CFO)',
    'Capital Expenditures (Capex)',
    'Net Cash Flow',
    'Total Cash',
    'Total Debt',
    'Working Capital',
    'Current Ratio',
    'Quick Ratio',
    'Equity (Book Value)',
    'Debt-to-Equity Ratio',
    'Debt-to-Assets Ratio',
    'Return on Equity (ROE %)',
    'Return on Assets (ROA %)',
    'Asset Turnover',
    'Receivables Turnover',
    'Days Sales Outstanding (DSO)',
    'Inventory Turnover',
    'Interest Coverage Ratio',
    'Debt/EBITDA',
    'Number of Employees',
]

def get_api_key():
    """Return Companies House API key from CH_API_KEY environment variable."""
    key = os.environ.get('CH_API_KEY', '').strip()
    if not key:
        raise EnvironmentError(
            "CH_API_KEY environment variable is not set. "
            "Set it to your Companies House API key before running, e.g.:\n"
            "  export CH_API_KEY=your-key-here  (Linux/macOS)\n"
            "  set CH_API_KEY=your-key-here     (Windows CMD)\n"
            "  $env:CH_API_KEY='your-key-here'  (PowerShell)"
        )
    return key


def request_with_retries(url, auth=None, params=None, headers=None, timeout=30, retries=3):
    """Make an HTTP GET request with retry support, handling both network errors and 429 rate limits."""
    for attempt in range(retries):
        try:
            response = requests.get(url, auth=auth, params=params, headers=headers, timeout=timeout)
            if response.status_code == 429:
                wait = int(response.headers.get('Retry-After', 2 ** (attempt + 1)))
                logger.warning("Rate limited (429) for %s — retrying in %ds...", url, wait)
                time.sleep(wait)
                continue
            return response
        except requests.exceptions.RequestException as e:
            if attempt == retries - 1:
                logger.error("Network request failed for %s: %s", url, e)
                return None
            time.sleep(1)
    logger.error("All %d retries exhausted for %s", retries, url)
    return None


def get_company_name(company_number, api_key):
    """Return the company name for display/export purposes."""
    try:
        response = request_with_retries(f"{BASE_URL}/company/{company_number}", auth=(api_key, ''))
        if response and response.status_code == 200:
            return response.json().get('company_name', company_number)
    except Exception:
        pass
    return company_number


def search_companies_by_name(name, api_key, max_results=10):
    """Search Companies House by name. Returns list of dicts with 'number', 'name', 'status'."""
    response = request_with_retries(
        f"{BASE_URL}/search/companies",
        params={'q': name, 'items_per_page': max_results},
        auth=(api_key, '')
    )
    if not response or response.status_code != 200:
        return []
    return [
        {
            'number': item['company_number'],
            'name': item.get('title', 'Unknown'),
            'status': item.get('company_status', ''),
        }
        for item in response.json().get('items', [])
    ]


def get_company_input():
    """Prompt user for company name or number"""
    while True:
        user_input = input("Enter company name or company number: ").strip()
        if user_input:
            return user_input
        print("Input cannot be empty. Please try again.")

def find_company_number(name_or_number, api_key):
    """Find company number from name or validate number"""
    # Check if it's a number (any length)
    if name_or_number.isdigit():
        # Pad to 8 digits with leading zeros
        padded_number = name_or_number.zfill(8)
        try:
            response = request_with_retries(f"{BASE_URL}/company/{padded_number}", auth=(api_key, ''))
            if response.status_code == 200:
                company_data = response.json()
                print(f"Found company: {company_data.get('company_name', 'Unknown')}")
                return padded_number
            elif response.status_code == 401:
                print(f"ERROR: API authentication failed (401).")
                print(f"Please verify your API key is valid and activated.")
                print(f"Response: {response.json().get('error', 'Unknown error')}")
                return None
            elif response.status_code == 404:
                print(f"Company number {padded_number} not found.")
                return None
            else:
                print(f"Error: {response.status_code} - {response.json().get('error', response.text[:100])}")
                return None
        except Exception as e:
            print(f"Error validating company number: {e}")
            return None
    else:
        # Search by name - show multiple results
        try:
            response = request_with_retries(f"{BASE_URL}/search/companies",
                                             params={'q': name_or_number, 'items_per_page': 10},
                                             auth=(api_key, ''))
            if response.status_code == 200:
                data = response.json()
                items = data.get('items', [])
                if items:
                    print(f"\nFound {data.get('total_count', len(items))} companies. Showing first 10:")
                    for i, company in enumerate(items, 1):
                        print(f"  {i}. {company.get('title', 'Unknown')} ({company['company_number']})")
                    
                    choice = input("\nEnter the number of the correct company (or 0 to cancel): ").strip()
                    if choice.isdigit() and 1 <= int(choice) <= len(items):
                        return items[int(choice) - 1]['company_number']
                    else:
                        print("Invalid selection.")
                        return None
                else:
                    print(f"No companies found matching '{name_or_number}'.")
                    return None
            elif response.status_code == 401:
                print(f"ERROR: API authentication failed (401).")
                print(f"Reason: {response.json().get('error', 'Unknown error')}")
                print(f"\nPlease verify:")
                print(f"  - Your API key is valid and not expired")
                print(f"  - Your API key is activated in the Companies House developer portal")
                print(f"  - Your IP address is not blocked")
                return None
            else:
                print(f"Search failed: {response.status_code} - {response.json().get('error', response.text[:100])}")
                return None
        except Exception as e:
            print(f"Error searching for company: {e}")
            return None

def get_accounts_filings(company_number, api_key, years=10):
    """Get list of accounts filings for the last N years"""
    try:
        filings = []
        start_index = 0
        page_size = 100
        cutoff_date = datetime.now() - timedelta(days=years*365)

        while True:
            response = request_with_retries(
                f"{BASE_URL}/company/{company_number}/filing-history",
                auth=(api_key, ''),
                params={'items_per_page': page_size, 'start_index': start_index}
            )
            if not response:
                break
            if response.status_code == 401:
                print(f"ERROR: API authentication failed (401). Cannot retrieve filing history.")
                return []
            elif response.status_code != 200:
                print(f"Failed to get filing history: {response.status_code} - {response.json().get('error', response.text[:100])}")
                return []

            page_data = response.json()
            page_items = page_data.get('items', [])
            if not page_items:
                break

            filings.extend(page_items)
            if len(page_items) < page_size:
                break
            start_index += page_size

        accounts_filings = []
        for filing in filings:
            if filing.get('type') == 'AA' and filing.get('date'):
                filing_date = datetime.strptime(filing['date'], '%Y-%m-%d')
                if filing_date >= cutoff_date:
                    accounts_filings.append(filing)
        return sorted(accounts_filings, key=lambda x: datetime.strptime(x['date'], '%Y-%m-%d'))
    except Exception as e:
        logger.error("Error getting filing history: %s", e, exc_info=True)
        return []

def extract_xbrl_values(xbrl_content, filing_date):
    """Extract financial values from iXBRL document"""
    try:
        # Parse XML
        try:
            root = etree.fromstring(xbrl_content)
        except etree.XMLSyntaxError:
            # If strict parsing fails, retry with error recovery
            parser = etree.XMLParser(recover=True)
            root = etree.fromstring(xbrl_content, parser=parser)
        
        extracted_data = {'date': filing_date}
        contexts = extract_xbrl_contexts(root)
        concept_facts = {}
        
        # Extract all numeric or tagged iXBRL elements
        for elem in root.iter():
            tag = elem.tag if isinstance(elem.tag, str) else str(elem.tag)
            local_tag = tag.split('}')[-1] if '}' in tag else tag
            text = (elem.text or '').strip() if elem.text else ''
            if not text:
                continue
            
            # Skip common HTML/XHTML structural tags by local tag name only
            if local_tag.lower() in {'html', 'head', 'body', 'div', 'span', 'p', 'br', 'script', 'style', 'title', 'meta', 'link'}:
                continue
            
            # Identify the concept name using the XBRL element name or tag name
            elem_name = elem.get('name', '') or ''
            if ':' in elem_name:
                concept = elem_name.split(':')[-1]
            else:
                concept = elem_name or local_tag
            concept = concept.strip()
            if not concept:
                continue
            
            # Parse numeric values from text content broadly
            value = parse_numeric_text(text)
            if value is not None:
                concept_lower = concept.lower()
                context_ref = elem.get('contextRef')
                fact = build_numeric_fact(concept, value, elem, context_ref, contexts, filing_date)
                concept_facts.setdefault(concept_lower, []).append(fact)
                continue
            
            # Preserve non-numeric XBRL flags and labels
            if 'nonnumeric' in tag.lower() or 'nonfraction' in tag.lower() or elem_name:
                concept_lower = concept.lower()
                extracted_data[concept] = text
                extracted_data[concept_lower] = text

        for concept_lower, facts in concept_facts.items():
            best_fact = select_best_fact(facts, filing_date, concept_lower)
            if best_fact is None:
                continue
            concept_name = best_fact['concept']
            extracted_data[concept_lower] = best_fact['value']
            extracted_data[concept_name] = best_fact['value']

        extracted_data['__contexts__'] = contexts
        extracted_data['__facts__'] = concept_facts
        
        return extracted_data if len(extracted_data) > 1 else None
        
    except Exception as e:
        logger.error("Error extracting XBRL values: %s", e, exc_info=True)
        return None


def extract_xbrl_contexts(root):
    """Extract context metadata used to choose the most relevant fact for the filing period."""
    contexts = {}
    for elem in root.iter():
        tag = elem.tag if isinstance(elem.tag, str) else str(elem.tag)
        local_tag = tag.split('}')[-1].lower() if '}' in tag else str(tag).lower()
        if local_tag != 'context':
            continue

        context_id = elem.get('id')
        if not context_id:
            continue

        context_info = {
            'id': context_id,
            'startdate': None,
            'enddate': None,
            'instant': None,
            'dimensions': 0,
        }

        for child in elem.iter():
            child_tag = child.tag if isinstance(child.tag, str) else str(child.tag)
            child_local = child_tag.split('}')[-1].lower() if '}' in child_tag else str(child_tag).lower()
            child_text = (child.text or '').strip() if child.text else ''
            if child_local == 'startdate' and child_text:
                context_info['startdate'] = child_text
            elif child_local == 'enddate' and child_text:
                context_info['enddate'] = child_text
            elif child_local == 'instant' and child_text:
                context_info['instant'] = child_text
            elif child_local in {'explicitmember', 'typedmember'}:
                context_info['dimensions'] += 1

        contexts[context_id] = context_info
    return contexts


def build_numeric_fact(concept, value, elem, context_ref, contexts, filing_date):
    """Normalize a numeric fact, including sign and scale metadata."""
    normalized_value = apply_fact_attributes(value, elem)
    return {
        'concept': concept,
        'value': normalized_value,
        'context_ref': context_ref,
        'context': contexts.get(context_ref, {}),
        'filing_date': filing_date,
    }


def apply_fact_attributes(value, elem):
    """Apply iXBRL sign and scale attributes to a parsed numeric fact."""
    normalized_value = value
    scale = elem.get('scale')
    sign = (elem.get('sign') or '').strip()

    if scale not in (None, ''):
        try:
            normalized_value = normalized_value * (10 ** int(scale))
        except (TypeError, ValueError, OverflowError):
            pass

    if sign == '-':
        normalized_value = -abs(normalized_value)

    return normalized_value


def select_best_fact(facts, filing_date, concept_name=None):
    """Choose the fact most likely to represent the current reporting period."""
    if not facts:
        return None

    filing_dt = safe_parse_date(filing_date)
    statement_type = classify_concept_statement_type(concept_name)

    def score_fact(fact):
        score = 0
        context = fact.get('context') or {}
        end_dt = safe_parse_date(context.get('enddate'))
        instant_dt = safe_parse_date(context.get('instant'))
        reference_dt = end_dt or instant_dt

        if filing_dt and reference_dt:
            day_diff = abs((filing_dt - reference_dt).days)
            score += max(0, SCORE_DATE_PROXIMITY_MAX - min(day_diff, SCORE_DATE_PROXIMITY_MAX))
            if day_diff == 0:
                score += SCORE_DATE_EXACT_BONUS

        if statement_type in {'income_statement', 'cash_flow'}:
            if context.get('enddate') and context.get('startdate'):
                score += SCORE_DURATION_CONTEXT_BONUS
            if context.get('instant'):
                score -= SCORE_INSTANT_CONTEXT_PENALTY
        elif statement_type == 'balance_sheet':
            if context.get('instant'):
                score += SCORE_INSTANT_CONTEXT_BONUS
            if context.get('enddate') and context.get('startdate'):
                score -= SCORE_DURATION_CONTEXT_PENALTY
        else:
            if context.get('enddate'):
                score += SCORE_UNKNOWN_DURATION_BONUS
            if context.get('instant'):
                score += SCORE_UNKNOWN_INSTANT_BONUS

        dimensions = context.get('dimensions', 0)
        if dimensions == 0:
            score += SCORE_NO_DIMENSION_BONUS
        else:
            score -= dimensions * SCORE_DIMENSION_PENALTY

        if fact.get('value') not in (None, 0):
            score += 5

        return score

    best_fact = max(facts, key=score_fact)
    best_fact = dict(best_fact)
    best_fact['selection_score'] = score_fact(best_fact)
    return best_fact


def classify_concept_statement_type(concept_name):
    """Infer whether a concept belongs to the income statement, balance sheet, or cash flow."""
    concept = (concept_name or '').lower()
    if not concept:
        return 'unknown'

    for statement_type, keywords in STATEMENT_TYPE_KEYWORDS.items():
        if concept in keywords:
            return statement_type
    for statement_type, keywords in STATEMENT_TYPE_KEYWORDS.items():
        if any(keyword in concept for keyword in keywords if len(keyword) >= 4):
            return statement_type
    return 'unknown'


def safe_parse_date(value):
    """Parse a YYYY-MM-DD date safely."""
    if not value:
        return None
    try:
        return datetime.strptime(str(value), '%Y-%m-%d')
    except (TypeError, ValueError):
        return None


def parse_numeric_text(text):
    """Parse a formatted numeric string like '18,076,799' into a float."""
    if not text:
        return None
    cleaned = str(text).strip()
    cleaned = cleaned.replace('£', '').replace('$', '').replace('€', '')
    cleaned = cleaned.replace(',', '').replace('\u00A0', '').replace('\u2013', '-').replace('\u2212', '-')
    cleaned = cleaned.replace('(', '-').replace(')', '').strip().lower()
    if not cleaned or cleaned in ['-', 'nil', 'n/a']:
        return None
    multiplier = 1.0
    if cleaned.endswith('billion'):
        multiplier = 1_000_000_000
        cleaned = cleaned[:-7].strip()
    elif cleaned.endswith('million'):
        multiplier = 1_000_000
        cleaned = cleaned[:-7].strip()
    elif cleaned.endswith('bn'):
        multiplier = 1_000_000_000
        cleaned = cleaned[:-2].strip()
    elif cleaned.endswith('m'):
        multiplier = 1_000_000
        cleaned = cleaned[:-1].strip()
    elif cleaned.endswith('k'):
        multiplier = 1_000
        cleaned = cleaned[:-1].strip()
    cleaned = cleaned.strip()

    if re.fullmatch(r'-?\d+(?:\.\d+)?', cleaned):
        try:
            return float(cleaned) * multiplier
        except ValueError:
            return None
    return None


def ensure_metric_details(metrics):
    """Create the internal metric-details container if needed."""
    return metrics.setdefault('__metric_details__', {})


def confidence_label(score):
    """Map a numeric confidence score onto a readable label."""
    if score >= CONFIDENCE_HIGH_THRESHOLD:
        return 'high'
    if score >= CONFIDENCE_MEDIUM_THRESHOLD:
        return 'medium'
    if score >= CONFIDENCE_LOW_THRESHOLD:
        return 'low'
    return 'very low'


def set_metric_detail(metrics, metric, source, confidence, concept=None, detail=None):
    """Store provenance for a metric without affecting the numeric output shape."""
    details = ensure_metric_details(metrics)
    details[metric] = {
        'source': source,
        'confidence': max(0, min(100, int(round(confidence)))),
        'confidence_label': confidence_label(confidence),
        'concept': concept,
        'detail': detail,
    }


def append_warning(metrics, message):
    """Append a warning message without duplicating existing text."""
    if not message:
        return
    existing = str(metrics.get('Warning', '') or '').strip()
    if message in existing:
        return
    metrics['Warning'] = f"{existing} {message}".strip() if existing else message


def get_metric_detail(metrics, metric):
    """Fetch stored provenance for a metric, if present."""
    return metrics.get('__metric_details__', {}).get(metric)


def derive_metric_confidence(metrics, dependencies, base_confidence=72):
    """Estimate derived-metric confidence from the weakest input confidence."""
    dependency_scores = []
    for dependency in dependencies:
        dependency_detail = get_metric_detail(metrics, dependency)
        if dependency_detail:
            dependency_scores.append(dependency_detail.get('confidence', base_confidence))
    if not dependency_scores:
        return base_confidence
    return max(35, min(dependency_scores) - 5)


def should_accept_metric_override(metrics, metric, candidate_value, candidate_source, candidate_confidence, tolerance_ratio=0.05):
    """Decide whether a new candidate should replace an existing metric value."""
    existing_value = metrics.get(metric)
    existing_detail = get_metric_detail(metrics, metric)
    if existing_value is None or existing_detail is None:
        return True

    existing_confidence = existing_detail.get('confidence', 0)
    existing_source = existing_detail.get('source', 'unknown')
    if existing_value == candidate_value:
        return candidate_confidence >= existing_confidence

    scale = max(abs(float(existing_value)) if isinstance(existing_value, (int, float, np.floating)) else 0, 1)
    difference_ratio = abs(float(candidate_value) - float(existing_value)) / scale if isinstance(candidate_value, (int, float, np.floating)) and isinstance(existing_value, (int, float, np.floating)) else 1

    if existing_source == 'ixbrl fact' and existing_confidence >= candidate_confidence + 10:
        append_warning(
            metrics,
            f"Retained {metric} from stronger {existing_source} evidence instead of weaker {candidate_source}."
        )
        return False

    if existing_confidence >= candidate_confidence and difference_ratio <= tolerance_ratio:
        return False

    return candidate_confidence >= existing_confidence or difference_ratio > 0.25


def set_derived_metric_detail(metrics, metric, dependencies, detail=None):
    """Record provenance for a derived metric."""
    confidence = derive_metric_confidence(metrics, dependencies)
    dep_text = ', '.join(dependencies)
    detail_text = detail or f"Derived from {dep_text}."
    set_metric_detail(metrics, metric, 'derived', confidence, detail=detail_text)


def summarize_metric_details(metrics):
    """Create a concise, readable extraction summary for notes/comments."""
    summaries = []
    for metric in METRIC_SUMMARY_ORDER:
        if metrics.get(metric) is None:
            continue
        metric_detail = get_metric_detail(metrics, metric)
        if not metric_detail:
            continue
        summary = f"{metric}: {metric_detail['source']} ({metric_detail['confidence_label']})"
        if metric_detail.get('concept'):
            summary += f" [{metric_detail['concept']}]"
        elif metric_detail.get('detail'):
            summary += f" [{metric_detail['detail']}]"
        summaries.append(summary)
    return 'Extraction summary: ' + '; '.join(summaries) + '.' if summaries else ''


def finalize_metric_metadata(metrics):
    """Persist a text summary of provenance for export while keeping internal details hidden."""
    metrics['Extraction Summary'] = summarize_metric_details(metrics)
    return metrics


def build_notes_text(metrics):
    """Combine warnings and extraction provenance into a single note/comment string."""
    notes = []
    warning_value = metrics.get('Warning', '')
    summary_value = metrics.get('Extraction Summary', '')
    warning_text = '' if pd.isna(warning_value) else str(warning_value or '').strip()
    summary_text = '' if pd.isna(summary_value) else str(summary_value or '').strip()
    if warning_text:
        notes.append(warning_text)
    if summary_text:
        notes.append(summary_text)
    return ' '.join(notes).strip()


def validate_metric_consistency(metrics):
    """Flag obvious cross-metric inconsistencies and down-rank dubious derived values."""
    revenue = metrics.get('Revenue')
    gross_profit = metrics.get('Gross Profit')
    ebit = metrics.get('Operating Income (EBIT)')
    ebitda = metrics.get('EBITDA')
    depreciation = metrics.get('Depreciation and Amortisation')
    total_debt = metrics.get('Total Debt')
    total_cash = metrics.get('Total Cash')
    net_debt = metrics.get('Net Debt')

    if revenue not in (None, 0) and gross_profit is not None and gross_profit > revenue * 1.05:
        append_warning(metrics, 'Validation warning: Gross Profit exceeds Revenue by more than 5%; review the selected values.')

    if ebit is not None and ebitda is not None and depreciation not in (None, 0) and ebitda + 1 < ebit:
        append_warning(metrics, 'Validation warning: EBITDA is below EBIT despite positive depreciation/amortisation; review source selection.')
        ebitda_detail = get_metric_detail(metrics, 'EBITDA')
        if ebitda_detail:
            ebitda_detail['confidence'] = max(25, ebitda_detail['confidence'] - 20)
            ebitda_detail['confidence_label'] = confidence_label(ebitda_detail['confidence'])

    for margin_metric in ('Gross Margin (%)', 'Operating Margin (%)', 'EBITDA Margin (%)', 'Net Margin (%)'):
        margin_value = metrics.get(margin_metric)
        if margin_value is not None and abs(float(margin_value)) > 100:
            append_warning(metrics, f'Validation warning: {margin_metric} is outside +/-100%; review source selection.')

    if total_debt is not None and total_cash is not None and net_debt is not None:
        expected_net_debt = total_debt - total_cash
        if abs(float(net_debt) - float(expected_net_debt)) > max(1, abs(expected_net_debt) * 0.02):
            append_warning(metrics, 'Validation warning: Net Debt does not reconcile to Total Debt minus Total Cash.')

    if metrics.get('Cash Flow from Operations (CFO)') is not None and metrics.get('Net Income') not in (None, 0):
        conversion = metrics.get('Cash Flow Conversion (CFO ÷ Net Income)')
        if conversion is not None and abs(float(conversion)) > 10:
            append_warning(metrics, 'Validation warning: Cash Flow Conversion is unusually high in magnitude; review CFO and Net Income sources.')


def fact_confidence_from_selection(selection_score, dimensions):
    """Convert fact-selection score into a confidence score."""
    base_score = CONFIDENCE_FACT_BASE + min(max(selection_score, 0), CONFIDENCE_SCORE_CAP) * CONFIDENCE_SCORE_WEIGHT
    if dimensions == 0:
        base_score += CONFIDENCE_NO_DIMENSION_BONUS
    elif dimensions > 1:
        base_score -= min(dimensions * CONFIDENCE_DIMENSION_PENALTY, CONFIDENCE_DIMENSION_PENALTY_CAP)
    return max(CONFIDENCE_MINIMUM, min(CONFIDENCE_MAXIMUM, int(round(base_score))))


def describe_fact_detail(fact):
    """Build a short human-readable descriptor for a chosen XBRL fact."""
    context = fact.get('context') or {}
    if context.get('enddate') and context.get('startdate'):
        return f"duration context {context['startdate']} to {context['enddate']}"
    if context.get('instant'):
        return f"instant context {context['instant']}"
    if fact.get('context_ref'):
        return f"context {fact['context_ref']}"
    return 'best-matching fact'


def get_xbrl_fact(extracted_data, keys, excluded_substrings=None):
    """Return the best matching fact object for the requested concepts."""
    excluded_substrings = excluded_substrings or []
    fact_map = extracted_data.get('__facts__', {}) if isinstance(extracted_data, dict) else {}
    if not fact_map:
        return None

    key_set = [k.lower() for k in keys if k]
    for key in key_set:
        facts = fact_map.get(key)
        if facts:
            return select_best_fact(facts, extracted_data.get('date'), key)

    generic_terms = {'income', 'profit', 'cash', 'tax', 'costs', 'interest'}
    best_match = None
    best_match_score = -1
    for concept, facts in fact_map.items():
        if any(excl in concept for excl in excluded_substrings):
            continue
        for key in key_set:
            if key not in concept:
                continue
            if key in generic_terms and concept != key:
                continue
            match_score = 0
            if concept == key:
                match_score += 100
            if concept.startswith(key) or concept.endswith(key):
                match_score += 20
            match_score += min(len(key), 30)
            best_fact = select_best_fact(facts, extracted_data.get('date'), concept)
            total_score = match_score + best_fact.get('selection_score', 0)
            if total_score > best_match_score:
                best_match_score = total_score
                best_match = best_fact

    return best_match


def extract_narrative_metric(xhtml_content, metric, return_raw=False):
    """Extract a financial metric from narrative commentary when XBRL tags are missing."""
    patterns_map = {
        'Revenue': [
            r'\b(?:sales|turnover|revenue)\s+(?:was|is|of|for|during|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:sales|turnover|revenue)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*(?:sales|turnover|revenue)\b',
        ],
        'Gross Profit': [
            r'\bgross profit\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\bgross profit\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*gross profit\b',
        ],
        'Operating Income (EBIT)': [
            r'\b(?:operating profit|operating income|ebit)\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:operating profit|operating income|ebit)\s*(?:\d{1,3}\s*)?£?([\d,]{2,}(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*(?:operating profit|operating income|ebit)\b',
        ],
        'EBITDA': [
            r'\b(?:ebitda|earnings before interest depreciation and amortisation|earnings before interest depreciation & amortisation|earnings before interest and tax depreciation amortisation)\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:ebitda|earnings before interest depreciation and amortisation|earnings before interest depreciation & amortisation|earnings before interest and tax depreciation amortisation)\s*£?([\d,]{2,}(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*ebitda\b',
        ],
        'Net Income': [
            r'\b(?:profit for the year|profit after tax|profit after taxation|net profit|net income|profit after tax)\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:profit for the year|profit after tax|profit after taxation|net profit|net income)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*(?:profit for the year|profit after tax|profit after taxation|net profit|net income)\b',
        ],
        'Cash Flow from Operations (CFO)': [
            r'\b(?:cash generated from operations|net cash from operating activities|cash flow from operations|cash flow from operating activities)\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:cash generated from operations|net cash generated from operating activities|net cash from operating activities|cash flow from operations|cash flow from operating activities)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*(?:cash generated from operations|net cash from operating activities|cash flow from operations|cash flow from operating activities)\b',
        ],
        'Capital Expenditures (Capex)': [
            r'\b(?:capital expenditure|capital expenditures|capex)\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:capital expenditure|capital expenditures|capex)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b£([\d,]+(?:\.\d+)?)\s*(?:capital expenditure|capital expenditures|capex)\b',
        ],
        'Exceptionals': [
            r'\b(?:exceptional items|exceptional costs|exceptional cost|non-recurring items|one-off costs|one off costs)\s+(?:was|is|of|amounted to|amounting to|at)\s*£?([\d,]{5,}(?:\.\d+)?)\b',
            r'\b(?:exceptional items|exceptional costs|exceptional cost|non-recurring items|one-off costs|one off costs)\s*£?([\d,]{5,}(?:\.\d+)?)\b',
            r'\b£([\d,]{5,}(?:\.\d+)?)\s*(?:exceptional items|exceptional costs|exceptional cost|non-recurring items|one-off costs|one off costs)\b',
            r'\b(?:exceptional items|exceptional costs|exceptional cost)\b.{0,200}?\bcharge for the year\b.{0,80}?£?([\d,]{5,}(?:\.\d+)?)\b',
            r'\bcharge for the year\b.{0,120}?\b(?:exceptional items|exceptional costs|exceptional cost)\b.{0,80}?£?([\d,]{5,}(?:\.\d+)?)\b',
            r'\b(?:exceptional items|exceptional costs|exceptional cost)\b.{0,120}?£?([\d,]{5,}(?:\.\d+)?)\b',
        ],
        'Depreciation and Amortisation': [
            r'\b(?:depreciation and amortisation|depreciation & amortisation)\s+(?:was|is|of|for|amounted to|amounting to|at)\s*£?([\d,]+(?:\.\d+)?)\b',
            r'\b(?:depreciation and amortisation|depreciation & amortisation)\s*£?([\d,]+(?:\.\d+)?)\b',
        ],
        'Number of Employees': [
            r'\b(?:average number of(?: group)? employees(?: during (?:the )?(?:year|period))?)\s+(?:was|is|of|for)\s*([\d,]{1,7})\b',
            r'\b(?:average number of(?: group)? employees(?: during (?:the )?(?:year|period))?)\s*([\d,]{1,7})\b',
            r'\b(?:group|consolidated)\s+employees\s+(?:was|is|of|for)\s*([\d,]{1,7})\b',
        ],
    }

    patterns = patterns_map.get(metric, [])
    if not patterns:
        return None

    search_texts = build_document_search_texts(xhtml_content)

    for search_text in search_texts:
        for pattern in patterns:
            for match in re.finditer(pattern, search_text, flags=re.IGNORECASE):
                amount = parse_numeric_text(match.group(1))
                if amount is not None:
                    if return_raw:
                        return amount, match.group(0)
                    return amount

    return (None, None) if return_raw else None


def build_document_search_texts(document_content):
    """Create normalized text variants from XHTML, XML, or plain extracted document text."""
    raw_text = document_content.decode('utf-8', errors='ignore') if isinstance(document_content, (bytes, bytearray)) else str(document_content or '')
    raw_text = raw_text.strip()
    if not raw_text:
        return []

    cleaned_text = re.sub(r'<[^>]+>', ' ', raw_text)
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
    search_texts = [cleaned_text] if cleaned_text else []

    if isinstance(document_content, (bytes, bytearray)):
        try:
            parser = etree.HTMLParser(recover=True)
            root = etree.fromstring(document_content, parser)
        except Exception:
            try:
                parser = etree.XMLParser(recover=True)
                root = etree.fromstring(document_content, parser)
            except Exception:
                root = None

        if root is not None:
            body_text = ' '.join(t.strip() for t in root.itertext() if t and t.strip())
            body_text = re.sub(r'\s+', ' ', body_text).strip()
            if body_text and body_text not in search_texts:
                search_texts.append(body_text)

    return search_texts


def extract_pdf_text(pdf_content):
    """Extract plain text from a PDF document for low-confidence fallback parsing."""
    if PdfReader is None:
        logger.warning('PDF text extraction requested but pypdf is not installed. Install pypdf to enable PDF fallback.')
        return None

    try:
        reader = PdfReader(BytesIO(pdf_content))
        page_text = []
        for page in reader.pages:
            text = page.extract_text() or ''
            if text.strip():
                page_text.append(text)
        return '\n'.join(page_text).strip() or None
    except Exception as e:
        logger.warning('PDF text extraction failed: %s', e)
        return None


def extract_narrative_da_components(xhtml_content):
    """Extract consolidated D&A components from narrative text or note disclosures."""
    search_texts = build_document_search_texts(xhtml_content)
    component_patterns = {
        'amortisation': [
            r'\bamortisation of intangible assets\b.{0,200}?\bcharge for the year\b.{0,60}?£?([\d,]+(?:\.\d+)?)\b',
            r'\bamortisation of intangible assets\b.{0,80}?£?([\d,]+(?:\.\d+)?)\b',
            r'\bcharge for the year\b.{0,120}?\bamortisation of intangible assets\b.{0,60}?£?([\d,]+(?:\.\d+)?)\b',
        ],
        'depreciation': [
            r'\bdepreciation of tangible assets\b.{0,200}?\bcharge for the year\b.{0,60}?£?([\d,]+(?:\.\d+)?)\b',
            r'\bdepreciation of tangible assets\b.{0,80}?£?([\d,]+(?:\.\d+)?)\b',
            r'\bcharge for the year\b.{0,120}?\bdepreciation of tangible assets\b.{0,60}?£?([\d,]+(?:\.\d+)?)\b',
        ],
    }

    components = {'amortisation': None, 'depreciation': None}
    for search_text in search_texts:
        for component, patterns in component_patterns.items():
            if components[component] is not None:
                continue
            for pattern in patterns:
                for match in re.finditer(pattern, search_text, flags=re.IGNORECASE):
                    amount = parse_numeric_text(match.group(1))
                    if amount is not None:
                        components[component] = amount
                        break
                if components[component] is not None:
                    break

    total = None
    if components['amortisation'] is not None or components['depreciation'] is not None:
        total = sum(value for value in components.values() if value is not None)

    return {
        'amortisation': components['amortisation'],
        'depreciation': components['depreciation'],
        'total': total,
    }


def calculate_pdf_fallback_metrics(pdf_text, filing_date):
    """Extract a limited set of metrics from PDF text when no structured iXBRL resource exists."""
    metrics = calculate_financial_metrics({'date': filing_date})
    pdf_warning = (
        'PDF fallback used because Companies House did not provide an XHTML/iXBRL resource for this filing. '
        'Figures below were extracted from unstructured PDF text and may be incomplete or less reliable than iXBRL-derived values.'
    )

    fallback_fields = [
        'Revenue',
        'Gross Profit',
        'Operating Income (EBIT)',
        'EBITDA',
        'Exceptionals',
        'Number of Employees',
        'Net Income',
        'Cash Flow from Operations (CFO)',
        'Capital Expenditures (Capex)'
    ]

    extracted_fields = []
    for field in fallback_fields:
        if field == 'Exceptionals':
            narrative_value, raw_match_text = extract_narrative_metric(pdf_text, field, return_raw=True)
        else:
            narrative_value = extract_narrative_metric(pdf_text, field)
            raw_match_text = None
        if narrative_value is None:
            continue
        if field == 'Exceptionals':
            if 1900 <= abs(float(narrative_value)) <= 2100 and raw_match_text and not re.search(r'£|gbp|pound|pounds', raw_match_text, flags=re.IGNORECASE):
                continue
            narrative_value = abs(float(narrative_value))
        if field == 'Number of Employees':
            narrative_value = int(round(float(narrative_value)))
        metrics[field] = narrative_value
        set_metric_detail(metrics, field, 'pdf narrative fallback', CONFIDENCE_PDF_NARRATIVE, detail='Unstructured PDF text match.')
        extracted_fields.append(field)

    narrative_da = extract_narrative_da_components(pdf_text)
    if narrative_da.get('total') is not None:
        metrics['Depreciation and Amortisation'] = narrative_da['total']
        set_metric_detail(metrics, 'Depreciation and Amortisation', 'pdf note disclosure', CONFIDENCE_PDF_NOTE, detail='Parsed from PDF note disclosure text.')
        da_parts = []
        if narrative_da.get('amortisation') is not None:
            da_parts.append(f"amortisation={narrative_da['amortisation']:,.0f}")
        if narrative_da.get('depreciation') is not None:
            da_parts.append(f"depreciation={narrative_da['depreciation']:,.0f}")
        extracted_fields.append('Depreciation and Amortisation')
        pdf_warning += ' PDF note-disclosure parsing was used for depreciation and amortisation'
        if da_parts:
            pdf_warning += f" ({', '.join(da_parts)})"
        pdf_warning += '.'

    revenue = metrics.get('Revenue')
    gross_profit = metrics.get('Gross Profit')
    operating_income = metrics.get('Operating Income (EBIT)')
    ebitda = metrics.get('EBITDA')
    exceptionals = metrics.get('Exceptionals')
    depreciation = metrics.get('Depreciation and Amortisation')
    net_income = metrics.get('Net Income')
    operating_cf = metrics.get('Cash Flow from Operations (CFO)')

    if ebitda is None and operating_income is not None and depreciation is not None:
        ebitda = operating_income + depreciation
        metrics['EBITDA'] = ebitda
        set_derived_metric_detail(metrics, 'EBITDA', ['Operating Income (EBIT)', 'Depreciation and Amortisation'])

    metrics['Gross Margin (%)'] = calculate_value(gross_profit, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
    metrics['Operating Margin (%)'] = calculate_value(operating_income, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
    metrics['EBITDA Margin (%)'] = calculate_value(ebitda, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
    metrics['Net Margin (%)'] = calculate_value(net_income, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
    metrics['Adjusted EBITDA'] = calculate_value(ebitda, exceptionals, lambda a, b: a + (b if b is not None else 0) if a is not None else None)
    metrics['Cash Flow Conversion (CFO ÷ Net Income)'] = calculate_value(operating_cf, net_income, lambda a, b: a/b if a is not None and b not in (0, None) else None)
    if metrics['Gross Margin (%)'] is not None:
        set_derived_metric_detail(metrics, 'Gross Margin (%)', ['Gross Profit', 'Revenue'])
    if metrics['Operating Margin (%)'] is not None:
        set_derived_metric_detail(metrics, 'Operating Margin (%)', ['Operating Income (EBIT)', 'Revenue'])
    if metrics['EBITDA Margin (%)'] is not None:
        set_derived_metric_detail(metrics, 'EBITDA Margin (%)', ['EBITDA', 'Revenue'])
    if metrics['Net Margin (%)'] is not None:
        set_derived_metric_detail(metrics, 'Net Margin (%)', ['Net Income', 'Revenue'])
    if metrics['Adjusted EBITDA'] is not None:
        set_derived_metric_detail(metrics, 'Adjusted EBITDA', ['EBITDA', 'Exceptionals'])
    if metrics['Cash Flow Conversion (CFO ÷ Net Income)'] is not None:
        set_derived_metric_detail(metrics, 'Cash Flow Conversion (CFO ÷ Net Income)', ['Cash Flow from Operations (CFO)', 'Net Income'])
    metrics['Warning'] = pdf_warning
    if extracted_fields:
        metrics['Warning'] += ' Extracted fields: ' + ', '.join(extracted_fields) + '.'
    else:
        metrics['Warning'] += ' No supported financial metrics were confidently identified in the PDF text.'

    metrics['Filing_Date'] = filing_date
    validate_metric_consistency(metrics)
    return finalize_metric_metadata(metrics)


def get_xbrl_value(data_lower, keys, excluded_substrings=None):
    """Search extracted data using exact-first matching with safer substring fallback."""
    excluded_substrings = excluded_substrings or []
    key_set = [k.lower() for k in keys if k]

    # 1) Exact match first
    for key in key_set:
        if key in data_lower:
            return data_lower[key]

    # 2) Safer substring match with scoring to avoid accidental concept collisions
    #    (e.g., matching 'income' inside 'deferredincome').
    generic_terms = {'income', 'profit', 'cash', 'tax', 'costs', 'interest'}
    best_candidate = None
    best_score = -1

    for concept, value in data_lower.items():
        if any(excl in concept for excl in excluded_substrings):
            continue

        for key in key_set:
            if key in concept:
                # Avoid broad generic substring matches unless exact.
                if key in generic_terms and concept != key:
                    continue

                score = 0
                if concept == key:
                    score += 100
                if concept.startswith(key) or concept.endswith(key):
                    score += 20
                score += min(len(key), 30)

                if score > best_score:
                    best_score = score
                    best_candidate = value

    return best_candidate


def get_group_employee_count(data_lower):
    """Prefer consolidated/group employee disclosures; fallback to best available employee fact."""
    consolidated_keys = [
        'consolidatedaveragenumberemployeesduringperiod',
        'groupaveragenumberemployeesduringperiod',
        'consolidatedaveragenumberemployees',
        'groupaveragenumberemployees',
    ]

    for key in consolidated_keys:
        if key in data_lower and data_lower[key] is not None:
            return data_lower[key]

    employee_candidates = []
    for concept, value in data_lower.items():
        if 'employ' not in concept:
            continue
        if not isinstance(value, (int, float, np.floating)):
            continue
        if value <= 0:
            continue
        employee_candidates.append((concept, value))

    if not employee_candidates:
        return None

    consolidated_candidates = [
        (concept, value)
        for concept, value in employee_candidates
        if ('group' in concept or 'consolidated' in concept)
    ]
    if consolidated_candidates:
        return max(consolidated_candidates, key=lambda item: item[1])[1]

    # Fallback: choose largest employee count concept (typically group over company-only).
    return max(employee_candidates, key=lambda item: item[1])[1]


def has_income_statement_data(data_lower):
    """Detect whether the document contains income statement-related concepts."""
    income_markers = [
        'turnover', 'revenue', 'sales', 'profitloss', 'netincome',
        'operatingprofit', 'profitbeforetax', 'profitaftertax'
    ]
    return any(any(marker in k for marker in income_markers) for k in data_lower)


def calculate_financial_metrics(extracted_data):
    """Calculate financial metrics from extracted XBRL data"""
    metrics = {}
    
    # Create lowercase lookup dict for case-insensitive matching
    data_lower = {k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}
    
    try:
        # Revenue/Turnover
        revenue_keys = XBRL_TAG_MAPPINGS.get('Revenue', []) + ['revenues']
        revenue_fact = get_xbrl_fact(
            extracted_data,
            revenue_keys,
            excluded_substrings=['deferredincome', 'accruedincome', 'comprehensiveincome', 'otherincome']
        )
        revenue = revenue_fact['value'] if revenue_fact else get_xbrl_value(
            data_lower,
            revenue_keys,
            excluded_substrings=['deferredincome', 'accruedincome', 'comprehensiveincome', 'otherincome']
        )
        metrics['Revenue'] = revenue
        if revenue_fact:
            set_metric_detail(metrics, 'Revenue', 'ixbrl fact', fact_confidence_from_selection(revenue_fact.get('selection_score', 0), (revenue_fact.get('context') or {}).get('dimensions', 0)), concept=revenue_fact.get('concept'), detail=describe_fact_detail(revenue_fact))

        # Cost of Sales / Gross Profit
        cost_keys = XBRL_TAG_MAPPINGS.get('CostOfSales', []) + ['costofsalesafterpurchasedgoods']
        cost_of_sales = get_xbrl_value(data_lower, cost_keys)
        gross_profit_keys = XBRL_TAG_MAPPINGS.get('GrossProfit', [])
        gross_profit_fact = get_xbrl_fact(extracted_data, gross_profit_keys)
        gross_profit = gross_profit_fact['value'] if gross_profit_fact else get_xbrl_value(data_lower, gross_profit_keys)
        if gross_profit is None and revenue is not None and cost_of_sales is not None:
            gross_profit = revenue - cost_of_sales
        metrics['Gross Profit'] = gross_profit
        if gross_profit_fact:
            set_metric_detail(metrics, 'Gross Profit', 'ixbrl fact', fact_confidence_from_selection(gross_profit_fact.get('selection_score', 0), (gross_profit_fact.get('context') or {}).get('dimensions', 0)), concept=gross_profit_fact.get('concept'), detail=describe_fact_detail(gross_profit_fact))
        elif gross_profit is not None and revenue is not None and cost_of_sales is not None:
            set_derived_metric_detail(metrics, 'Gross Profit', ['Revenue'])
        metrics['Gross Margin (%)'] = calculate_value(gross_profit, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Gross Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'Gross Margin (%)', ['Gross Profit', 'Revenue'])
        
        # Operating Income (EBIT)
        operating_keys = XBRL_TAG_MAPPINGS.get('OperatingIncome', [])
        operating_fact = get_xbrl_fact(
            extracted_data,
            operating_keys,
            excluded_substrings=['aftertax', 'forperiod']
        )
        operating_income = operating_fact['value'] if operating_fact else get_xbrl_value(
            data_lower,
            operating_keys,
            excluded_substrings=['aftertax', 'forperiod']
        )
        metrics['Operating Income (EBIT)'] = operating_income
        if operating_fact:
            set_metric_detail(metrics, 'Operating Income (EBIT)', 'ixbrl fact', fact_confidence_from_selection(operating_fact.get('selection_score', 0), (operating_fact.get('context') or {}).get('dimensions', 0)), concept=operating_fact.get('concept'), detail=describe_fact_detail(operating_fact))
        metrics['Operating Margin (%)'] = calculate_value(operating_income, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Operating Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'Operating Margin (%)', ['Operating Income (EBIT)', 'Revenue'])
        
        # EBITDA
        ebitda_keys = XBRL_TAG_MAPPINGS.get('EBITDA', []) + ['earningsbeforeinteresttaxdepreciationandamortisation', 'adjustedoperatingprofit']
        ebitda_fact = get_xbrl_fact(extracted_data, ebitda_keys)
        ebitda = ebitda_fact['value'] if ebitda_fact else get_xbrl_value(data_lower, ebitda_keys)
        # Try combined D&A concept first
        combined_da_keys = ['depreciationandamortisation', 'depreciationandamortisationexpense',
                            'depreciationandamortisationcharge', 'totalamortisationanddepreciation',
                            'depreciationamortisation']
        depreciation = get_xbrl_value(data_lower, combined_da_keys)

        if depreciation is None:
            # Sum separate depreciation and amortisation when no combined concept exists
            # Explicitly look for: amortisation of intangible assets + depreciation of tangible assets
            # Priority to P&L period costs (year charges) over accumulations or general terms
            dep_keys = ['consolidateddepreciationchargeforyearpropertyplantequipment',
                        'consolidatedincreasesfromdepreciationchargesforyearpropertyplantequipment',
                        'consolidatedusefullivesofproperty',
                        'increasefromdepreciationchargeforyearpropertyplantequipment',
                        'depreciationoftangibleassets', 'depreciationoftangiblefixedassets',
                        'depreciationcharge', 'depreciationexpense', 'depreciationtangibleassets',
                        'depreciationfixedassets']
            amort_keys = ['consolidatedamortisationofintangibleassets',
                          'consolidatedamortisationexpense',
                          'consolidatedamortisationofintangiblefixedassets',
                          'increasefromamortisationchargeforyearintangibleassets',
                          'amortisationexpense', 'amortisationofintangibleassets',
                          'amortisationofintangiblefixedassets', 'amortisationcharge',
                          'amortisationintangibleassets', 'amortisation',
                          'depreciationandamortisationexpense', 'totalamortisation',
                          'amortisationcosts', 'amortisationofintangibles']
            dep_only = get_xbrl_value(data_lower, dep_keys)
            amort_only = get_xbrl_value(data_lower, amort_keys)
            if dep_only is not None and amort_only is not None:
                depreciation = dep_only + amort_only
            elif dep_only is not None:
                depreciation = dep_only
            elif amort_only is not None:
                depreciation = amort_only

        if depreciation is None:
            # Broad fallback
            fallback_da_keys = XBRL_TAG_MAPPINGS.get('DepreciationAmortization', []) + [
                'depreciation', 'amortisation', 'depreciationexpense', 'amortisationexpense']
            depreciation = get_xbrl_value(data_lower, fallback_da_keys)

        exceptional_keys = XBRL_TAG_MAPPINGS.get('Exceptionals', []) + [
            'exceptionaladministrativecosts', 'exceptionaloperatingexpenses',
            'exceptionaloperatingcosts', 'nonrecurringitems', 'oneoffcosts'
        ]
        exceptional_fact = get_xbrl_fact(extracted_data, exceptional_keys)
        exceptionals = exceptional_fact['value'] if exceptional_fact else get_xbrl_value(data_lower, exceptional_keys)
        if exceptionals is not None:
            exceptionals = abs(float(exceptionals))
            if 1900 <= exceptionals <= 2100:
                exceptionals = None

        if ebitda is None and operating_income is not None and depreciation is not None:
            ebitda = operating_income + depreciation
        metrics['EBITDA'] = ebitda
        if ebitda_fact:
            set_metric_detail(metrics, 'EBITDA', 'ixbrl fact', fact_confidence_from_selection(ebitda_fact.get('selection_score', 0), (ebitda_fact.get('context') or {}).get('dimensions', 0)), concept=ebitda_fact.get('concept'), detail=describe_fact_detail(ebitda_fact))
        elif ebitda is not None and operating_income is not None and depreciation is not None:
            set_derived_metric_detail(metrics, 'EBITDA', ['Operating Income (EBIT)', 'Depreciation and Amortisation'])
        metrics['Exceptionals'] = exceptionals
        if exceptional_fact and exceptionals is not None:
            set_metric_detail(metrics, 'Exceptionals', 'ixbrl fact', fact_confidence_from_selection(exceptional_fact.get('selection_score', 0), (exceptional_fact.get('context') or {}).get('dimensions', 0)), concept=exceptional_fact.get('concept'), detail=describe_fact_detail(exceptional_fact))
        metrics['Adjusted EBITDA'] = calculate_value(ebitda, exceptionals, lambda a, b: a + (b if b is not None else 0) if a is not None else None)
        if metrics['Adjusted EBITDA'] is not None:
            set_derived_metric_detail(metrics, 'Adjusted EBITDA', ['EBITDA', 'Exceptionals'])
        metrics['Depreciation and Amortisation'] = depreciation
        if depreciation is not None:
            set_metric_detail(metrics, 'Depreciation and Amortisation', 'ixbrl fact or composite', 68 if combined_da_keys else 60, detail='Selected from depreciation/amortisation concepts.')
        metrics['EBITDA Margin (%)'] = calculate_value(ebitda, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['EBITDA Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'EBITDA Margin (%)', ['EBITDA', 'Revenue'])
        
        # Net Income (Profit for period)
        net_income_keys = XBRL_TAG_MAPPINGS.get('NetIncome', [])
        net_income_fact = get_xbrl_fact(extracted_data, net_income_keys)
        net_income = net_income_fact['value'] if net_income_fact else get_xbrl_value(data_lower, net_income_keys)
        metrics['Net Income'] = net_income
        if net_income_fact:
            set_metric_detail(metrics, 'Net Income', 'ixbrl fact', fact_confidence_from_selection(net_income_fact.get('selection_score', 0), (net_income_fact.get('context') or {}).get('dimensions', 0)), concept=net_income_fact.get('concept'), detail=describe_fact_detail(net_income_fact))
        metrics['Net Margin (%)'] = calculate_value(net_income, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Net Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'Net Margin (%)', ['Net Income', 'Revenue'])
        
        # EPS
        eps_keys = XBRL_TAG_MAPPINGS.get('EPS', [])
        eps_fact = get_xbrl_fact(extracted_data, eps_keys)
        metrics['Earnings Per Share (EPS)'] = eps_fact['value'] if eps_fact else get_xbrl_value(data_lower, eps_keys)
        if eps_fact:
            set_metric_detail(metrics, 'Earnings Per Share (EPS)', 'ixbrl fact', fact_confidence_from_selection(eps_fact.get('selection_score', 0), (eps_fact.get('context') or {}).get('dimensions', 0)), concept=eps_fact.get('concept'), detail=describe_fact_detail(eps_fact))
        
        # Balance Sheet Items
        current_assets_keys = XBRL_TAG_MAPPINGS.get('CurrentAssets', [])
        current_assets_fact = get_xbrl_fact(extracted_data, current_assets_keys)
        current_assets = current_assets_fact['value'] if current_assets_fact else get_xbrl_value(data_lower, current_assets_keys)
        total_assets_keys = XBRL_TAG_MAPPINGS.get('TotalAssets', []) + ['totalassetslesscurrentliabilities']
        total_assets_fact = get_xbrl_fact(extracted_data, total_assets_keys)
        total_assets = total_assets_fact['value'] if total_assets_fact else get_xbrl_value(data_lower, total_assets_keys)
        
        current_liabilities_keys = XBRL_TAG_MAPPINGS.get('CurrentLiabilities', []) + ['creditorsdue', 'tradecreditorstradepayables']
        current_liabilities_fact = get_xbrl_fact(extracted_data, current_liabilities_keys)
        current_liabilities = current_liabilities_fact['value'] if current_liabilities_fact else get_xbrl_value(data_lower, current_liabilities_keys)
        long_term_keys = XBRL_TAG_MAPPINGS.get('LongTermLiabilities', []) + ['longtermborrowings', 'longtermloans']
        long_term_debt_fact = get_xbrl_fact(extracted_data, long_term_keys)
        long_term_debt = long_term_debt_fact['value'] if long_term_debt_fact else get_xbrl_value(data_lower, long_term_keys)
        total_debt = calculate_value(current_liabilities, long_term_debt, lambda a, b: a + b if a is not None and b is not None else a if a is not None else b)
        
        equity_keys = XBRL_TAG_MAPPINGS.get('Equity', []) + ['shareholdersfunds']
        equity_fact = get_xbrl_fact(extracted_data, equity_keys)
        equity = equity_fact['value'] if equity_fact else get_xbrl_value(data_lower, equity_keys)
        inventory_keys = XBRL_TAG_MAPPINGS.get('Inventory', [])
        inventory_fact = get_xbrl_fact(extracted_data, inventory_keys)
        inventory = inventory_fact['value'] if inventory_fact else get_xbrl_value(data_lower, inventory_keys)
        cash_keys = XBRL_TAG_MAPPINGS.get('Cash', []) + ['cashandcashequivalents', 'cashatbank']
        cash_fact = get_xbrl_fact(extracted_data, cash_keys)
        cash = cash_fact['value'] if cash_fact else get_xbrl_value(data_lower, cash_keys)
        receivables_keys = XBRL_TAG_MAPPINGS.get('Receivables', []) + ['tradedebtorstradereceivables']
        receivables_fact = get_xbrl_fact(extracted_data, receivables_keys)
        receivables = receivables_fact['value'] if receivables_fact else get_xbrl_value(data_lower, receivables_keys)

        if current_assets_fact:
            set_metric_detail(metrics, 'Current Assets', 'ixbrl fact', fact_confidence_from_selection(current_assets_fact.get('selection_score', 0), (current_assets_fact.get('context') or {}).get('dimensions', 0)), concept=current_assets_fact.get('concept'), detail=describe_fact_detail(current_assets_fact))
        if total_assets_fact:
            set_metric_detail(metrics, 'Total Assets', 'ixbrl fact', fact_confidence_from_selection(total_assets_fact.get('selection_score', 0), (total_assets_fact.get('context') or {}).get('dimensions', 0)), concept=total_assets_fact.get('concept'), detail=describe_fact_detail(total_assets_fact))
        if current_liabilities_fact:
            set_metric_detail(metrics, 'Current Liabilities', 'ixbrl fact', fact_confidence_from_selection(current_liabilities_fact.get('selection_score', 0), (current_liabilities_fact.get('context') or {}).get('dimensions', 0)), concept=current_liabilities_fact.get('concept'), detail=describe_fact_detail(current_liabilities_fact))
        if long_term_debt_fact:
            set_metric_detail(metrics, 'Long-Term Debt', 'ixbrl fact', fact_confidence_from_selection(long_term_debt_fact.get('selection_score', 0), (long_term_debt_fact.get('context') or {}).get('dimensions', 0)), concept=long_term_debt_fact.get('concept'), detail=describe_fact_detail(long_term_debt_fact))
        if equity_fact:
            set_metric_detail(metrics, 'Equity (Book Value)', 'ixbrl fact', fact_confidence_from_selection(equity_fact.get('selection_score', 0), (equity_fact.get('context') or {}).get('dimensions', 0)), concept=equity_fact.get('concept'), detail=describe_fact_detail(equity_fact))
        if inventory_fact:
            set_metric_detail(metrics, 'Inventory', 'ixbrl fact', fact_confidence_from_selection(inventory_fact.get('selection_score', 0), (inventory_fact.get('context') or {}).get('dimensions', 0)), concept=inventory_fact.get('concept'), detail=describe_fact_detail(inventory_fact))
        if receivables_fact:
            set_metric_detail(metrics, 'Receivables', 'ixbrl fact', fact_confidence_from_selection(receivables_fact.get('selection_score', 0), (receivables_fact.get('context') or {}).get('dimensions', 0)), concept=receivables_fact.get('concept'), detail=describe_fact_detail(receivables_fact))
        
        # Working Capital and Ratios
        working_capital = calculate_value(current_assets, current_liabilities, lambda a, b: a - b if a is not None and b is not None else None)
        metrics['Working Capital'] = working_capital
        if working_capital is not None:
            set_derived_metric_detail(metrics, 'Working Capital', ['Current Assets', 'Current Liabilities'])
        metrics['Current Ratio'] = calculate_value(current_assets, current_liabilities, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Current Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Current Ratio', ['Current Assets', 'Current Liabilities'])
        quick_assets = calculate_value(current_assets, inventory, lambda a, b: a - b if a is not None and b is not None else a)
        if quick_assets is not None:
            set_derived_metric_detail(metrics, 'Quick Assets', ['Current Assets', 'Inventory'])
        metrics['Quick Ratio'] = calculate_value(quick_assets, current_liabilities, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Quick Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Quick Ratio', ['Quick Assets', 'Current Liabilities'])
        
        # Debt Metrics
        metrics['Total Debt'] = total_debt
        metrics['Total Cash'] = cash
        metrics['Net Debt'] = calculate_value(total_debt, cash, lambda a, b: a - b if a is not None and b is not None else None)
        if total_debt is not None:
            set_metric_detail(metrics, 'Total Debt', 'ixbrl fact or composite', 70, detail='Derived from current and long-term debt concepts.')
        if cash is not None:
            set_metric_detail(metrics, 'Total Cash', 'ixbrl fact', 72, detail='Selected from cash/cash equivalents concepts.')
        if metrics['Net Debt'] is not None:
            set_derived_metric_detail(metrics, 'Net Debt', ['Total Debt', 'Total Cash'])
        metrics['Debt-to-Equity Ratio'] = calculate_value(total_debt, equity, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Debt-to-Equity Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Debt-to-Equity Ratio', ['Total Debt', 'Equity (Book Value)'])
        metrics['Debt-to-Assets Ratio'] = calculate_value(total_debt, total_assets, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Debt-to-Assets Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Debt-to-Assets Ratio', ['Total Debt', 'Total Assets'])
        
        # Equity
        metrics['Equity (Book Value)'] = equity
        
        # Return Ratios
        metrics['Return on Equity (ROE %)'] = calculate_value(net_income, equity, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Return on Equity (ROE %)'] is not None:
            set_derived_metric_detail(metrics, 'Return on Equity (ROE %)', ['Net Income', 'Equity (Book Value)'])
        metrics['Return on Assets (ROA %)'] = calculate_value(net_income, total_assets, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Return on Assets (ROA %)'] is not None:
            set_derived_metric_detail(metrics, 'Return on Assets (ROA %)', ['Net Income', 'Total Assets'])
        metrics['Asset Turnover'] = calculate_value(revenue, total_assets, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Asset Turnover'] is not None:
            set_derived_metric_detail(metrics, 'Asset Turnover', ['Revenue', 'Total Assets'])
        
        # Working Capital Metrics
        metrics['Receivables Turnover'] = calculate_value(revenue, receivables, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Receivables Turnover'] is not None:
            set_derived_metric_detail(metrics, 'Receivables Turnover', ['Revenue', 'Receivables'])
        metrics['Days Sales Outstanding (DSO)'] = calculate_value(receivables, revenue, lambda a, b: (a/b*365) if a is not None and b not in (0, None) else None)
        if metrics['Days Sales Outstanding (DSO)'] is not None:
            set_derived_metric_detail(metrics, 'Days Sales Outstanding (DSO)', ['Receivables', 'Revenue'])
        metrics['Inventory Turnover'] = calculate_value(revenue, inventory, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Inventory Turnover'] is not None:
            set_derived_metric_detail(metrics, 'Inventory Turnover', ['Revenue', 'Inventory'])
        
        # Interest & Coverage
        interest_keys = XBRL_TAG_MAPPINGS.get('Interest', [])
        interest_fact = get_xbrl_fact(extracted_data, interest_keys)
        interest_expense = interest_fact['value'] if interest_fact else get_xbrl_value(data_lower, interest_keys)
        if interest_fact:
            set_metric_detail(metrics, 'Interest Expense', 'ixbrl fact', fact_confidence_from_selection(interest_fact.get('selection_score', 0), (interest_fact.get('context') or {}).get('dimensions', 0)), concept=interest_fact.get('concept'), detail=describe_fact_detail(interest_fact))
        metrics['Interest Coverage Ratio'] = calculate_value(operating_income, interest_expense, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Interest Coverage Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Interest Coverage Ratio', ['Operating Income (EBIT)', 'Interest Expense'])
        metrics['Debt/EBITDA'] = calculate_value(total_debt, ebitda, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Debt/EBITDA'] is not None:
            set_derived_metric_detail(metrics, 'Debt/EBITDA', ['Total Debt', 'EBITDA'])

        tax_keys = XBRL_TAG_MAPPINGS.get('Tax', []) + ['taxation', 'currenttax', 'corporationtax']
        tax_fact = get_xbrl_fact(extracted_data, tax_keys)
        metrics['Tax Paid'] = tax_fact['value'] if tax_fact else get_xbrl_value(data_lower, tax_keys)
        if tax_fact:
            set_metric_detail(metrics, 'Tax Paid', 'ixbrl fact', fact_confidence_from_selection(tax_fact.get('selection_score', 0), (tax_fact.get('context') or {}).get('dimensions', 0)), concept=tax_fact.get('concept'), detail=describe_fact_detail(tax_fact))

        metrics['Number of Employees'] = get_group_employee_count(data_lower)
        if metrics['Number of Employees'] is not None:
            set_metric_detail(metrics, 'Number of Employees', 'ixbrl fact', 70, detail='Best available group or consolidated employee concept.')
        
        # Cash Flow (if available in extended reports)
        operating_cf_keys = ['operatingcashflow', 'cashfromoperations', 'netcashfromoperatingactivities', 'cashgeneratedfromoperations', 'netcashgeneratedfromoperatingactivities']
        operating_cf_fact = get_xbrl_fact(extracted_data, operating_cf_keys)
        operating_cf = operating_cf_fact['value'] if operating_cf_fact else get_xbrl_value(data_lower, operating_cf_keys)
        metrics['Cash Flow from Operations (CFO)'] = operating_cf
        if operating_cf_fact:
            set_metric_detail(metrics, 'Cash Flow from Operations (CFO)', 'ixbrl fact', fact_confidence_from_selection(operating_cf_fact.get('selection_score', 0), (operating_cf_fact.get('context') or {}).get('dimensions', 0)), concept=operating_cf_fact.get('concept'), detail=describe_fact_detail(operating_cf_fact))
        capex_keys = [
            'capex',
            'capitalexpenditure',
            'purchaseoftangiblefixedassets',
            'purchaseofintangiblefixedassets',
            'purchaseofpropertyplantequipment',
            'purchaseofintangibleassets',
            'additionsotherthanthroughbusinesscombinationspropertyplantequipment',
            'additionsotherthanthroughbusinesscombinationsintangibleassets',
            'acquisitionoffixedassets',
            'acquisitionofpropertyplantequipment',
            'acquisitionofintangibleassets',
            'propertyplantequipment'
        ]
        capex_fact = get_xbrl_fact(extracted_data, capex_keys)
        capex = capex_fact['value'] if capex_fact else get_xbrl_value(data_lower, capex_keys)
        metrics['Capital Expenditures (Capex)'] = capex
        if capex_fact:
            set_metric_detail(metrics, 'Capital Expenditures (Capex)', 'ixbrl fact', fact_confidence_from_selection(capex_fact.get('selection_score', 0), (capex_fact.get('context') or {}).get('dimensions', 0)), concept=capex_fact.get('concept'), detail=describe_fact_detail(capex_fact))

        net_cash_flow_keys = [
            'netincreaseincashandcashequivalents',
            'increaseincashandcashequivalents',
            'decreaseincashandcashequivalents',
            'increaseDecreaseInCashAndCashEquivalents'.lower(),
            'netcashflow',
            'netcashinflowoutflow',
            'increaseDecreaseInCashBankOverdrafts'.lower(),
            'netincreaseDecreaseincashbankoverdrafts'.lower(),
        ]
        net_cash_fact = get_xbrl_fact(extracted_data, net_cash_flow_keys)
        net_cash_flow = net_cash_fact['value'] if net_cash_fact else get_xbrl_value(data_lower, net_cash_flow_keys)
        metrics['Net Cash Flow'] = net_cash_flow
        if net_cash_fact:
            set_metric_detail(metrics, 'Net Cash Flow', 'ixbrl fact', fact_confidence_from_selection(net_cash_fact.get('selection_score', 0), (net_cash_fact.get('context') or {}).get('dimensions', 0)), concept=net_cash_fact.get('concept'), detail=describe_fact_detail(net_cash_fact))
        metrics['Cash Flow Conversion (CFO ÷ Net Income)'] = calculate_value(operating_cf, net_income, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Cash Flow Conversion (CFO ÷ Net Income)'] is not None:
            set_derived_metric_detail(metrics, 'Cash Flow Conversion (CFO ÷ Net Income)', ['Cash Flow from Operations (CFO)', 'Net Income'])
        
        return metrics
    except Exception as e:
        logger.error("Error calculating financial metrics: %s", e, exc_info=True)
        return metrics

def calculate_value(val1, val2, operation):
    """Safely calculate value using operation, allowing None values"""
    try:
        return operation(val1, val2)
    except (TypeError, ValueError, ZeroDivisionError, ArithmeticError):
        pass
    return None


def format_financial_value(metric, value):
    """Format monetary values in millions with one decimal and brackets for negatives."""
    money_metrics = {
        'Revenue',
        'Gross Profit',
        'Operating Income (EBIT)',
        'EBITDA',
        'Net Income',
        'Working Capital',
        'Net Debt',
        'Equity (Book Value)',
        'Cash Flow from Operations (CFO)',
        'Net Cash Flow',
        'Capital Expenditures (Capex)'
    }
    if metric not in money_metrics or value is None:
        return None
    try:
        millions = float(value) / 1_000_000
        formatted = f"{abs(millions):.1f}"
        return f"({formatted})" if millions < 0 else formatted
    except Exception:
        return None

def download_and_parse_accounts(company_number, filing, api_key):
    """Download XBRL document from filing and extract financial data"""
    try:
        doc_metadata_link = filing.get('links', {}).get('document_metadata')
        if not doc_metadata_link:
            return None
        
        meta_response = request_with_retries(doc_metadata_link, auth=(api_key, ''))
        if not meta_response or meta_response.status_code != 200:
            return None
        
        metadata = meta_response.json()
        doc_link = metadata.get('links', {}).get('document')
        if not doc_link:
            return None

        resources = metadata.get('resources', {}) if isinstance(metadata.get('resources'), dict) else {}
        available_resource_types = {str(resource_type).lower() for resource_type in resources.keys()}
        has_structured_resource = any(
            resource_type in available_resource_types
            for resource_type in ('application/xhtml+xml', 'application/xml', 'text/html')
        )
        has_pdf_resource = 'application/pdf' in available_resource_types
        
        # Request XHTML format with Accept header
        # Note: doc_link already includes /content endpoint
        doc_response = request_with_retries(doc_link, 
                                            auth=(api_key, ''), 
                                            headers={'Accept': 'application/xhtml+xml'},
                                            timeout=30)
        if doc_response is not None and doc_response.status_code == 200:
            content_type = doc_response.headers.get('content-type', '')
            if 'pdf' in content_type.lower():
                doc_response = None
        elif doc_response is not None and doc_response.status_code == 406 and has_pdf_resource and not has_structured_resource:
            doc_response = None
        elif not doc_response or doc_response.status_code != 200:
            if has_pdf_resource:
                doc_response = None
            else:
                return None

        if doc_response is None and has_pdf_resource:
            pdf_warning_message = (
                'PDF fallback used because Companies House did not provide an XHTML/iXBRL resource for this filing. '
                'Figures may be incomplete or less reliable than structured XBRL extraction.'
            )
            logger.warning("PDF fallback for %s %s: %s", company_number, filing.get('date'), pdf_warning_message)
            pdf_response = request_with_retries(
                doc_link,
                auth=(api_key, ''),
                headers={'Accept': 'application/pdf'},
                timeout=30
            )
            if not pdf_response or pdf_response.status_code != 200:
                return {
                    'Warning': pdf_warning_message + ' PDF download failed, so no metrics could be extracted.',
                    'Extraction Summary': 'Extraction summary: PDF fallback failed before any metrics could be extracted.',
                    'Filing_Date': filing.get('date')
                }

            pdf_text = extract_pdf_text(pdf_response.content)
            if not pdf_text:
                return {
                    'Warning': pdf_warning_message + ' PDF text extraction failed, so no metrics could be extracted.',
                    'Extraction Summary': 'Extraction summary: PDF text extraction failed before any metrics could be extracted.',
                    'Filing_Date': filing.get('date')
                }

            return calculate_pdf_fallback_metrics(pdf_text, filing.get('date'))

        if doc_response is None:
            return None
        
        # Extract XBRL from iXBRL/XHTML
        extracted_data = extract_xbrl_values(doc_response.content, filing.get('date'))
        if not extracted_data or len(extracted_data) < 3:
            if has_pdf_resource:
                pdf_warning_message = (
                    'PDF fallback used because structured XBRL extraction did not return reliable metrics. '
                    'Figures may be incomplete or less reliable than structured XBRL extraction.'
                )
                logger.warning("PDF fallback for %s %s: %s", company_number, filing.get('date'), pdf_warning_message)
                pdf_response = request_with_retries(
                    doc_link,
                    auth=(api_key, ''),
                    headers={'Accept': 'application/pdf'},
                    timeout=30
                )
                if not pdf_response or pdf_response.status_code != 200:
                    return {
                        'Warning': pdf_warning_message + ' PDF download failed, so no metrics could be extracted.',
                        'Extraction Summary': 'Extraction summary: PDF fallback failed before any metrics could be extracted.',
                        'Filing_Date': filing.get('date')
                    }

                pdf_text = extract_pdf_text(pdf_response.content)
                if not pdf_text:
                    pdf_failure_reason = ' PDF text extraction failed before any metrics could be extracted.'
                    if PdfReader is None:
                        pdf_failure_reason = ' PDF text extraction is unavailable because pypdf is not installed.'
                    return {
                        'Warning': pdf_warning_message + pdf_failure_reason,
                        'Extraction Summary': 'Extraction summary: PDF text extraction failed before any metrics could be extracted.',
                        'Filing_Date': filing.get('date')
                    }

                return calculate_pdf_fallback_metrics(pdf_text, filing.get('date'))
            return None
        
        # Warn if the filing is an abbreviated account without a P&L
        warning_message = None
        abbrev_flag = extracted_data.get('entityhastakenexemptionundercompaniesactinnotpublishingitsownprofitlossaccounttruefalse')
        report_flag = extracted_data.get('reportincludesdetailedprofitlossstatementtruefalse')
        if abbrev_flag and str(abbrev_flag).lower() == 'true':
            warning_message = 'This filing has taken exemption from publishing a profit & loss account; many income statement metrics are unavailable.'
        elif report_flag and str(report_flag).lower() == 'false':
            warning_message = 'This filing does not include a detailed profit & loss statement; many income statement metrics are unavailable.'
        else:
            numeric_data = {k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}
            if not has_income_statement_data(numeric_data):
                warning_message = 'This filing does not contain identifiable income statement concepts in the XBRL extract; revenue and profit metrics may not be available.'

        if warning_message:
            logger.warning("Filing %s %s: %s", company_number, filing.get('date'), warning_message)

        # Calculate metrics
        metrics = calculate_financial_metrics(extracted_data)

        # Narrative fallback for missing key metrics
        fallback_fields = [
            'Revenue',
            'Gross Profit',
            'Operating Income (EBIT)',
            'EBITDA',
            'Exceptionals',
            'Number of Employees',
            'Net Income',
            'Cash Flow from Operations (CFO)',
            'Capital Expenditures (Capex)'
        ]
        fallback_warnings = []
        for field in fallback_fields:
            if metrics.get(field) is None:
                narrative_value = extract_narrative_metric(doc_response.content, field)
                if narrative_value is not None:
                    if field == 'Exceptionals':
                        narrative_value = abs(float(narrative_value))
                        if 1900 <= narrative_value <= 2100:
                            continue
                    if field == 'Number of Employees':
                        narrative_value = int(round(float(narrative_value)))
                    metrics[field] = narrative_value
                    set_metric_detail(metrics, field, 'xhtml narrative fallback', 58, detail='Narrative commentary fallback from XHTML/iXBRL filing text.')
                    fallback_warnings.append(f"{field} extracted from narrative commentary fallback.")

        # Employee-specific safeguard: if a narrative group/consolidated headcount appears
        # and is larger than the extracted XBRL figure, prefer it.
        narrative_employees = extract_narrative_metric(doc_response.content, 'Number of Employees')
        if narrative_employees is not None:
            narrative_employees = int(round(float(narrative_employees)))
            existing_employees = metrics.get('Number of Employees')
            if (
                (existing_employees is None or narrative_employees > existing_employees)
                and should_accept_metric_override(metrics, 'Number of Employees', narrative_employees, 'xhtml narrative fallback', 62)
            ):
                metrics['Number of Employees'] = narrative_employees
                set_metric_detail(metrics, 'Number of Employees', 'xhtml narrative fallback', 62, detail='Upgraded from narrative group disclosure fallback.')
                fallback_warnings.append('Number of Employees upgraded from narrative group disclosure fallback.')

        narrative_da = extract_narrative_da_components(doc_response.content)
        narrative_da_total = narrative_da.get('total')
        used_narrative_da = False
        existing_da = metrics.get('Depreciation and Amortisation')
        if narrative_da_total is not None and (
            existing_da is None or (existing_da not in (0, None) and narrative_da_total > existing_da * 1.5)
        ) and should_accept_metric_override(metrics, 'Depreciation and Amortisation', narrative_da_total, 'xhtml note disclosure', 68):
            metrics['Depreciation and Amortisation'] = narrative_da_total
            set_metric_detail(metrics, 'Depreciation and Amortisation', 'xhtml note disclosure', 68, detail='Narrative note disclosure override from XHTML/iXBRL filing text.')
            used_narrative_da = True
            da_parts = []
            if narrative_da.get('amortisation') is not None:
                da_parts.append(f"amortisation={narrative_da['amortisation']:,.0f}")
            if narrative_da.get('depreciation') is not None:
                da_parts.append(f"depreciation={narrative_da['depreciation']:,.0f}")
            detail_text = f" ({', '.join(da_parts)})" if da_parts else ''
            fallback_warnings.append(
                f"Depreciation and Amortisation extracted from narrative note disclosure{detail_text}."
            )

        if fallback_warnings:
            warning_text = ' '.join(fallback_warnings)
            metrics['Warning'] = ((warning_message + ' ') if warning_message else '') + warning_text

        # Recalculate derived metrics after narrative fallback values are applied
        revenue = metrics.get('Revenue')
        net_income = metrics.get('Net Income')
        gross_profit = metrics.get('Gross Profit')
        operating_income = metrics.get('Operating Income (EBIT)')
        ebitda = metrics.get('EBITDA')
        exceptionals = metrics.get('Exceptionals')
        depreciation = metrics.get('Depreciation and Amortisation')
        # Recalculate EBITDA if it was not set earlier or D&A was upgraded from narrative disclosure.
        if (ebitda is None or used_narrative_da) and operating_income is not None and depreciation is not None:
            ebitda = operating_income + depreciation
            metrics['EBITDA'] = ebitda
            set_derived_metric_detail(metrics, 'EBITDA', ['Operating Income (EBIT)', 'Depreciation and Amortisation'])
        metrics['Adjusted EBITDA'] = calculate_value(ebitda, exceptionals, lambda a, b: a + (b if b is not None else 0) if a is not None else None)
        if metrics['Adjusted EBITDA'] is not None:
            set_derived_metric_detail(metrics, 'Adjusted EBITDA', ['EBITDA', 'Exceptionals'])
        operating_cf = metrics.get('Cash Flow from Operations (CFO)')
        capex = metrics.get('Capital Expenditures (Capex)')
        net_cash_flow = metrics.get('Net Cash Flow')
        total_assets = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['totalassets', 'balancesheettotal', 'totalassetslesscurrentliabilities'])
        equity = metrics.get('Equity (Book Value)')
        receivables = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['tradedebtorstradereceivables', 'debtors', 'receivables'])
        inventory = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['inventory', 'stock'])
        current_assets = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['currentassets', 'currentassetslessstock', 'currentassetsincluding'])
        current_liabilities = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['creditorsduewithinoneyear', 'creditorsdue', 'creditors', 'tradecreditorstradepayables', 'tradecreditors', 'tradepayables'])
        long_term_debt = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['creditorsdueafteroneyear', 'longtermliabilities', 'longtermdebt', 'longtermborrowings', 'longtermloans'])
        total_debt = calculate_value(current_liabilities, long_term_debt, lambda a, b: a + b if a is not None and b is not None else a if a is not None else b)
        cash = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, ['cashandcashequivalents', 'cashbankonhand', 'cashcashequivalents', 'cash'])

        metrics['Gross Margin (%)'] = calculate_value(gross_profit, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Gross Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'Gross Margin (%)', ['Gross Profit', 'Revenue'])
        metrics['Operating Margin (%)'] = calculate_value(operating_income, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Operating Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'Operating Margin (%)', ['Operating Income (EBIT)', 'Revenue'])
        metrics['EBITDA Margin (%)'] = calculate_value(ebitda, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['EBITDA Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'EBITDA Margin (%)', ['EBITDA', 'Revenue'])
        metrics['Net Margin (%)'] = calculate_value(net_income, revenue, lambda a, b: (a/b*100) if a is not None and b not in (0, None) else None)
        if metrics['Net Margin (%)'] is not None:
            set_derived_metric_detail(metrics, 'Net Margin (%)', ['Net Income', 'Revenue'])
        metrics['Asset Turnover'] = calculate_value(revenue, total_assets, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Asset Turnover'] is not None:
            set_derived_metric_detail(metrics, 'Asset Turnover', ['Revenue', 'Total Assets'])
        metrics['Receivables Turnover'] = calculate_value(revenue, receivables, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Receivables Turnover'] is not None:
            set_derived_metric_detail(metrics, 'Receivables Turnover', ['Revenue', 'Receivables'])
        metrics['Days Sales Outstanding (DSO)'] = calculate_value(receivables, revenue, lambda a, b: (a/b*365) if a is not None and b not in (0, None) else None)
        if metrics['Days Sales Outstanding (DSO)'] is not None:
            set_derived_metric_detail(metrics, 'Days Sales Outstanding (DSO)', ['Receivables', 'Revenue'])
        metrics['Inventory Turnover'] = calculate_value(revenue, inventory, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Inventory Turnover'] is not None:
            set_derived_metric_detail(metrics, 'Inventory Turnover', ['Revenue', 'Inventory'])
        metrics['Debt-to-Equity Ratio'] = calculate_value(total_debt, equity, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Debt-to-Equity Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Debt-to-Equity Ratio', ['Total Debt', 'Equity (Book Value)'])
        metrics['Debt-to-Assets Ratio'] = calculate_value(total_debt, total_assets, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Debt-to-Assets Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Debt-to-Assets Ratio', ['Total Debt', 'Total Assets'])
        interest_value = get_xbrl_value({k.lower(): v for k, v in extracted_data.items() if k != 'date' and isinstance(v, (int, float))}, XBRL_TAG_MAPPINGS.get('Interest', []))
        metrics['Interest Coverage Ratio'] = calculate_value(operating_income, interest_value, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Interest Coverage Ratio'] is not None:
            set_derived_metric_detail(metrics, 'Interest Coverage Ratio', ['Operating Income (EBIT)', 'Interest Expense'])
        metrics['Debt/EBITDA'] = calculate_value(total_debt, ebitda, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Debt/EBITDA'] is not None:
            set_derived_metric_detail(metrics, 'Debt/EBITDA', ['Total Debt', 'EBITDA'])
        metrics['Net Cash Flow'] = net_cash_flow
        metrics['Cash Flow Conversion (CFO ÷ Net Income)'] = calculate_value(operating_cf, net_income, lambda a, b: a/b if a is not None and b not in (0, None) else None)
        if metrics['Cash Flow Conversion (CFO ÷ Net Income)'] is not None:
            set_derived_metric_detail(metrics, 'Cash Flow Conversion (CFO ÷ Net Income)', ['Cash Flow from Operations (CFO)', 'Net Income'])

        metrics['Warning'] = metrics.get('Warning', '')
        metrics['Filing_Date'] = filing.get('date')
        validate_metric_consistency(metrics)
        
        return finalize_metric_metadata(metrics)
        
    except Exception as e:
        logger.error("Error downloading/parsing accounts for %s %s: %s", company_number, filing.get('date'), e, exc_info=True)
        return None

def build_excel_workbook(all_metrics, company_name):
    """Build the Excel workbook from a list of per-filing metric dicts.

    Returns the workbook as raw bytes (suitable for writing to a file or
    returning directly from a web handler).
    """
    export_metrics = [
        {k: v for k, v in m.items() if not str(k).startswith('__')}
        for m in all_metrics
    ]
    df_raw = pd.DataFrame(export_metrics)
    df_raw['Filing_Date'] = pd.to_datetime(df_raw['Filing_Date'])
    df_raw = df_raw.sort_values('Filing_Date')
    df_raw['Year'] = df_raw['Filing_Date'].dt.year

    year_end_labels = {}
    for _, filing_row in df_raw.iterrows():
        year_key = str(filing_row['Year'])
        filing_date = filing_row.get('Filing_Date')
        if pd.notna(filing_date):
            year_end_labels[year_key] = pd.to_datetime(filing_date).strftime('%b-%y')

    duplicate_years = df_raw[df_raw.duplicated(subset='Year', keep=False)]['Year'].unique()
    for dup_year in duplicate_years:
        logger.warning("Multiple filings found for year %s; keeping the most recent one.", dup_year)

    template_years_df = df_raw.drop_duplicates(subset='Year', keep='last').copy()
    template_years_df['Revenue Growth (%)'] = template_years_df['Revenue'].pct_change() * 100 if 'Revenue' in template_years_df else np.nan
    template_years_df['Working Capital Movement'] = template_years_df['Working Capital'].diff() if 'Working Capital' in template_years_df else np.nan
    template_years_df['EBIT'] = template_years_df.get('Operating Income (EBIT)')
    template_years_df['EBIT Margin (%)'] = template_years_df.get('Operating Margin (%)')
    template_years_df['Exceptionals'] = template_years_df.get('Exceptionals')
    template_years_df['Adjusted EBITDA'] = template_years_df.apply(
        lambda row: row.get('EBITDA') + (row.get('Exceptionals') if pd.notna(row.get('Exceptionals')) else 0)
        if pd.notna(row.get('EBITDA')) else np.nan,
        axis=1
    )
    template_years_df['Capital Expenditures'] = template_years_df.get('Capital Expenditures (Capex)')
    template_years_df['Cash Flow from Operations'] = template_years_df.get('Cash Flow from Operations (CFO)')
    template_years_df['Net Cash Flow'] = template_years_df.get('Net Cash Flow')
    if 'Total Cash' in template_years_df and template_years_df['Net Cash Flow'].isna().all():
        template_years_df['Net Cash Flow'] = template_years_df['Total Cash'].diff()
    template_years_df['Notes:'] = template_years_df.apply(build_notes_text, axis=1)
    template_years_df = template_years_df.tail(TEMPLATE_MAX_YEARS).copy()

    wb = load_workbook(TEMPLATE_FILE)
    ws_template = wb[TEMPLATE_SHEET_NAME]

    if str(ws_template.cell(row=14, column=2).value or '').strip() == 'Depreciation and Amortisation':
        ws_template.insert_rows(14, amount=2)

    style_reference_row = 16
    for target_row, label in ((14, 'Exceptionals'), (15, 'Adjusted EBITDA')):
        for col_idx in range(1, ws_template.max_column + 1):
            ws_template.cell(row=target_row, column=col_idx)._style = copy(ws_template.cell(row=style_reference_row, column=col_idx)._style)
        ws_template.row_dimensions[target_row].height = ws_template.row_dimensions[style_reference_row].height
        ws_template.cell(row=target_row, column=2, value=label)

    ws_template.cell(row=13, column=2, value='')
    ws_template.cell(row=16, column=2, value='')

    for col_idx in range(1, ws_template.max_column + 1):
        ws_template.cell(row=17, column=col_idx)._style = copy(ws_template.cell(row=style_reference_row, column=col_idx)._style)
    ws_template.row_dimensions[17].height = ws_template.row_dimensions[style_reference_row].height
    ws_template.cell(row=17, column=2, value='Depreciation and Amortisation')
    ws_template.cell(row=TEMPLATE_ROW_MAP['Net Cash Flow'], column=2, value='Net Cash Flow')

    for col_offset in range(TEMPLATE_MAX_YEARS):
        col_idx = TEMPLATE_START_COL + col_offset
        ws_template.cell(row=2, column=col_idx, value='')
        ws_template.cell(row=3, column=col_idx, value='')
        for row_idx in TEMPLATE_ROW_MAP.values():
            ws_template.cell(row=row_idx, column=col_idx, value='')

    ws_template.cell(row=2, column=2, value=company_name)

    no_border = Border()
    table_last_col = TEMPLATE_START_COL + TEMPLATE_MAX_YEARS - 1
    for row_idx in range(2, TEMPLATE_ROW_MAP['Notes:'] + 1):
        for col_idx in range(2, table_last_col + 1):
            cell = ws_template.cell(row=row_idx, column=col_idx)
            cell.border = no_border
            fill_rgb = str(cell.fill.start_color.rgb).upper() if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb else ''
            is_blue_bar = cell.fill and cell.fill.fill_type == 'solid' and fill_rgb in TEMPLATE_BLUE_FILLS
            if not is_blue_bar:
                cell.fill = WHITE_FILL

    for col_offset, (_, filing_row) in enumerate(template_years_df.iterrows()):
        col_idx = TEMPLATE_START_COL + col_offset
        year_key = str(filing_row['Year'])
        year_cell = ws_template.cell(row=2, column=col_idx, value=year_key)
        year_end_cell = ws_template.cell(row=3, column=col_idx, value=year_end_labels.get(year_key, ''))
        year_cell.border = no_border
        year_end_cell.border = no_border

        for row_label, row_idx in TEMPLATE_ROW_MAP.items():
            value = filing_row.get(row_label)
            cell = ws_template.cell(row=row_idx, column=col_idx)
            cell.border = no_border
            cell.comment = None
            if value is None or pd.isna(value):
                cell.value = ''
            elif row_label in MONEY_TEMPLATE_ROWS:
                cell.value = round(float(value) / 1_000_000, 2)
                cell.number_format = '0.00;[Red](0.00)'
            elif row_label in PERCENT_TEMPLATE_ROWS:
                cell.value = round(float(value), 1)
                cell.number_format = '0.0'
            elif row_label in COUNT_TEMPLATE_ROWS:
                cell.value = int(round(float(value)))
                cell.number_format = '0'
            elif row_label == 'Notes:':
                note_text = str(value).strip()
                cell.value = 'See comment' if note_text else ''
                if note_text:
                    cell.comment = Comment(note_text, 'CH Agent')
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            else:
                cell.value = round(float(value), 2) if isinstance(value, (int, float, np.floating)) else value
                if isinstance(value, (int, float, np.floating)):
                    cell.number_format = '0.00'

            if row_label != 'Notes:' and cell.value != '':
                cell.alignment = Alignment(horizontal='right', vertical='center')

    ws_template.row_dimensions[TEMPLATE_ROW_MAP['Notes:']].height = 18

    if 'Raw Extracted Data' in wb.sheetnames:
        del wb['Raw Extracted Data']
    ws_raw = wb.create_sheet('Raw Extracted Data')
    for row in dataframe_to_rows(df_raw, index=False, header=True):
        ws_raw.append(row)

    header_fill = PatternFill(fill_type='solid', start_color='1F3A5F', end_color='1F3A5F')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center')
    body_font = Font(name='Calibri', size=10)
    ws_raw.freeze_panes = 'A2'
    ws_raw.sheet_view.showGridLines = False
    ws_raw.row_dimensions[1].height = 22
    for cell in ws_raw[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = no_border

    for col_idx in range(1, ws_raw.max_column + 1):
        col_letter = ws_raw.cell(row=1, column=col_idx).column_letter
        max_len = 0
        for row_idx in range(1, ws_raw.max_row + 1):
            cell_val = ws_raw.cell(row=row_idx, column=col_idx).value
            cell_len = len(str(cell_val)) if cell_val is not None else 0
            if cell_len > max_len:
                max_len = cell_len
        ws_raw.column_dimensions[col_letter].width = min(max(12, max_len + 2), 28)

    for row in ws_raw.iter_rows(min_row=2, max_row=ws_raw.max_row, min_col=1, max_col=ws_raw.max_column):
        for cell in row:
            cell.font = body_font
            cell.fill = WHITE_FILL
            if isinstance(cell.value, (int, float, np.floating)):
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = no_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def build_preview_df(all_metrics):
    """Build a concise year-by-year summary table for display in the web UI.

    Monetary values are expressed in £m; percentages and employee counts as-is.
    Returns a DataFrame with metrics as rows and filing years as columns.
    """
    PREVIEW_METRICS = [
        ('Revenue',             'money'),
        ('Gross Profit',        'money'),
        ('Gross Margin (%)',    'pct'),
        ('EBITDA',              'money'),
        ('EBITDA Margin (%)',   'pct'),
        ('Net Income',          'money'),
        ('Net Debt',            'money'),
        ('Number of Employees', 'count'),
    ]

    sorted_metrics = sorted(
        all_metrics,
        key=lambda m: pd.to_datetime(m.get('Filing_Date') or '1900-01-01')
    )

    # Build display column labels (e.g. "Apr-24"), deduplicating by year (keep last)
    seen_years = {}
    for m in sorted_metrics:
        dt = pd.to_datetime(m.get('Filing_Date') or '')
        label = dt.strftime('%b-%y') if pd.notna(dt) else '?'
        year = dt.year if pd.notna(dt) else 0
        seen_years[year] = (label, m)
    deduped = list(seen_years.values())[-TEMPLATE_MAX_YEARS:]
    col_labels = [label for label, _ in deduped]
    col_metrics = [m for _, m in deduped]

    rows = {}
    for metric_name, fmt in PREVIEW_METRICS:
        row = {}
        for col, m in zip(col_labels, col_metrics):
            val = m.get(metric_name)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                row[col] = None
            elif fmt == 'money':
                row[col] = round(float(val) / 1_000_000, 1)
            elif fmt == 'pct':
                row[col] = round(float(val), 1)
            elif fmt == 'count':
                row[col] = int(round(float(val)))
            else:
                row[col] = val
        rows[metric_name] = row

    df = pd.DataFrame(rows).T
    df.index.name = None
    return df


def run_analysis(company_number, api_key, on_progress=None):
    """Run the full filing retrieval and extraction pipeline for a company.

    Filings are downloaded in parallel (up to 5 concurrent requests) for speed.

    Args:
        company_number: 8-digit Companies House company number.
        api_key: Companies House API key.
        on_progress: optional callable(current, total, filing_date) called after
                     each filing completes (may be called from a worker thread).

    Returns:
        (excel_bytes, company_name, years_processed, preview_df, warnings).
        excel_bytes is None if no data could be extracted.
        preview_df is a summary DataFrame for display; warnings is a list of strings.
    """
    company_name = get_company_name(company_number, api_key)
    filings = get_accounts_filings(company_number, api_key, years=10)
    if not filings:
        return None, company_name, 0, None, []

    total = len(filings)
    completed = [0]
    lock = threading.Lock()

    def _fetch(filing):
        result = download_and_parse_accounts(company_number, filing, api_key)
        with lock:
            completed[0] += 1
            count = completed[0]
        if on_progress:
            on_progress(count, total, filing.get('date', ''))
        return result

    with ThreadPoolExecutor(max_workers=5) as executor:
        results = list(executor.map(_fetch, filings))

    all_metrics = [m for m in results if m is not None]
    if not all_metrics:
        return None, company_name, 0, None, []

    warnings = [
        f"{m.get('Filing_Date', 'Unknown date')}: {m['Warning'].strip()}"
        for m in all_metrics
        if m.get('Warning', '').strip()
    ]
    preview_df = build_preview_df(all_metrics)
    excel_bytes = build_excel_workbook(all_metrics, company_name)
    return excel_bytes, company_name, len(all_metrics), preview_df, warnings


def main():
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
    try:
        api_key = get_api_key()
        company_input = get_company_input()

        company_number = find_company_number(company_input, api_key)
        if not company_number:
            print("Could not find or validate company. Exiting.")
            return

        print(f"\nRetrieving accounts for company {company_number}...")

        filings = get_accounts_filings(company_number, api_key, years=10)
        if not filings:
            print("No accounts filings found in the last 10 years.")
            return
        print(f"Found {len(filings)} accounts filings.")

        def on_progress(current, total, filing_date):
            print(f"Processing accounts for {filing_date} ({current}/{total})...")

        excel_bytes, company_name, years, *_ = run_analysis(
            company_number, api_key, on_progress=on_progress
        )
        if not excel_bytes:
            print("No financial data could be extracted from filings.")
            return

        output_file = f"{company_number}_financial_analysis.xlsx"
        with open(output_file, 'wb') as f:
            f.write(excel_bytes)

        print(f"\nDone. Financial analysis for {company_name} exported to {output_file}")
        print(f"Successfully processed {years} years of financial documents")

    except Exception as e:
        print(f"An error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()